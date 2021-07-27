# -*- coding: utf-8 -*-
# 1.
# 修改日期 2021-06-24 修改人:蒙海全
# 修改排名排序问题 分场景合格率总计改为全区
# 2.
# 修改日期 2021-06-29 修改人:蒙海全
# 修改排名排序问题 千兆未归档七日均数据,千兆退单分析,新增核减后在途单，压单比采用新算法
# 3.
# 修改日期 2021-07-01 修改人:蒙海全
# 优化匹配网格方法,减少程序运用时间
# 4.
# 修改日期 2021-07-023 修改人:蒙海全
# 优化读取表格速度,减少程序运用时间

import pandas as pd
import xlwings as xw
from pandas import Series
import os
import datetime, time
from openpyxl import load_workbook
from openpyxl.styles import Font  # 字体样式
from openpyxl.styles import PatternFill  # 填充颜色
from openpyxl.styles import Alignment  # 字体对齐方式
from openpyxl.styles import Side, Border  # 边框样式

start_time = time.time()
# 日期调整
# today=datetime.datetime.now()  #系统取当天日期
today = datetime.datetime.strptime('2021-7-20', '%Y-%m-%d')  # 手动输入日期
today = datetime.datetime(today.year, today.month, today.day, 00, 00, 00)  # 将今日时间调整为当天0点
today1 = str(today.month) + '月' + today.strftime('%d') + '日'  # 日期转换str：9月10日

Y_day = today - datetime.timedelta(days=1)  # 昨天日期
before_yesterday = Y_day - datetime.timedelta(days=1)  # 前天日期

yesterday = Y_day.strftime('%Y') + '年' + str(Y_day.month) + '月' + Y_day.strftime('%d') + '日'  # 日期转换str
yesterday1 = str(Y_day.month) + '月' + Y_day.strftime('%d') + '日'  # 日期转换str

first_day = datetime.datetime(Y_day.year, Y_day.month, 1, 00, 00, 00)  # datetime类型 2019-09-01 00:00:00 当月1号开始
first_day1 = str(first_day.month) + '月' + first_day.strftime('%d') + '日'  # 日期转换str：9月01日

chao72 = today - datetime.timedelta(days=3)  # 超72小时
chao7day = today - datetime.timedelta(days=7)  # 超7天
chao30day = today - datetime.timedelta(days=30)  # 超30天
chao8 = datetime.datetime(Y_day.year, Y_day.month, Y_day.day, 16, 00, 00)  # 计算昨日16点 千兆工单即报即装临界点
chao8work = datetime.datetime(Y_day.year, Y_day.month, Y_day.day, 12, 00, 00)  # 计算昨日12点 千兆工单8工作时临界点
yesterday20 = datetime.datetime(Y_day.year, Y_day.month, Y_day.day, 20, 00, 00)  # 计算昨日20点
seven_day = today - datetime.timedelta(days=7)  # 七天前 用来计算7天日均归档量 datetime 类型


def GetDataFrame(file):
    if os.path.exists(file):
        try:
            app = xw.App(visible=False, add_book=False)
            wb = app.books.open(file)  # 打开Excel文件

            sht_all = wb.sheets[0]

            info = sht_all.used_range

            nrows = info.last_cell.row

            ncols = info.last_cell.column
            index1 = sht_all.range((1, 1), (1, ncols)).value
            index2 = Series(index1)
            Data = sht_all.range((2, 1), (nrows, ncols)).value
            Data = pd.DataFrame(Data, columns=index2)
            wb.close()
            app.quit()
            return Data
        except:
            # 报错自动结束进程
            print('运行错误')
            os.system('taskkill/IM et.exe /F')
            quit()
    else:
        print(f'文件名不存在--{file}')
        quit()


# ######################################################日期计算完毕
# 批量修改表格格式
def format_str(data, name, ty=str):
    data_types_dict = {name: ty}
    data = data.astype(data_types_dict)
    return data


# 计算总数函数
def data_total(data, cols, name, off=1):
    if off == 1:
        number = data.groupby([cols]).size().reset_index()
        number.rename(columns={0: '总数'}, inplace=True)
        number = number.append([{cols: '全区', '总数': number.apply(lambda x: x.sum()).总数}], ignore_index=True)
        number.rename(columns={'总数': name}, inplace=True)
    else:
        number = data.groupby([cols]).size().reset_index()
        number.rename(columns={0: name}, inplace=True)
    return number


# 多选项
def checkbox_total(data, col, name):
    quantity = data.groupby(['地市', col]).size().reset_index()
    quantity.rename(columns={0: '总数'}, inplace=True)
    quantity = quantity.append([{'地市': '全区', '总数': quantity.apply(lambda x: x.sum()).总数}], ignore_index=True)
    quantity.rename(columns={'总数': name}, inplace=True)
    return quantity


# 计算压单比
def total_ydb(data, daily, name):
    tp1 = data.drop(data.tail(1).index)  # 先删掉平均值那一行  ,inplace=True
    tp2 = tp1.iloc[:, 0:1]  # 日期列
    temp3 = tp1.iloc[:, 1:]  # 数据列
    temp4 = daily[[name, '未归档压单比']]  # .T.reset_index(drop=True)   #取出当日 日报 第八页计算好的 压单比
    temp4 = temp4[temp4[name].str.contains('代维信息为空') == False]
    temp4 = temp4.T.reset_index(drop=True)
    list_ = list(temp4.iloc[0])  # 取标段信息
    temp4.columns = list_  # 重命名 表头
    temp4 = temp4.drop(temp4.head(1).index)  # 删除多余的 那一行标段信息
    temp5 = pd.concat([temp3, temp4], axis=0)  # 拼接 今日的压单比
    temp5.loc['均值'] = temp5.apply(lambda x: x.mean())  # 列求均值
    temp5 = temp5.reset_index(drop=True)
    tp2 = tp2.append([{name: today1}], ignore_index=True)  # 添加日期
    tp2 = tp2.append([{name: '平均'}], ignore_index=True)  # 添加平均字段
    yd_new = pd.concat([tp2, temp5], axis=1)  # 拼接之前分开处理的汉字和数据
    return yd_new


def time_total(data, col):
    # data = data[data.是否是校园宽带 == '否']  # 本月不含校园工单数据
    # 全流程时长
    table = data[[col, '工单历时减去总时长分化时']].groupby([col]).mean().reset_index()
    table.rename(columns={'工单历时减去总时长分化时': '全流程时长'}, inplace=True)  # 改列名
    table['排名'] = table.rank(axis=0, ascending=True, method='dense').全流程时长  # 输出排名
    table = table.append(
        [{col: '全区', '全流程时长': data[[col, '工单历时减去总时长分化时']].mean().reset_index(name='全流程时长').at[0, '全流程时长'], '排名': ''}],
        ignore_index=True)  # 计算全区
    table = table.sort_values(by='全流程时长', ascending=True).reset_index(drop=True)  # 按时长 排序
    table = table.round({'全流程时长': 2})  # 四舍五入
    # 全流程及时率
    table_temp = data_total(data, col, '总工单数', off=1)
    table_temp = pd.merge(table_temp, data_total(data[data.是否超时 == '未超时'], col, '未超时工单数', off=1), on=col, how='left')
    table_temp['全流程及时率'] = (table_temp.未超时工单数 / table_temp.总工单数).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))  # 这里是计算各地市的及时率
    table_temp['排名'] = table_temp['全流程及时率'][:-1].rank(axis=0, ascending=False, method='dense')  # 输出排名
    table_temp = table_temp.sort_values(by='全流程及时率', ascending=False).reset_index(drop=True)  # 按时长 排序
    # 首次响应时长
    tp1 = data[[col, '首次响应时长']].groupby([col]).mean().reset_index()
    tp1['排名'] = tp1.rank(axis=0, ascending=True, method='dense').首次响应时长  # 输出排名
    tp1 = tp1.append(
        [{col: '全区', '首次响应时长': data[[col, '首次响应时长']].mean().reset_index(name='首次响应时长').at[0, '首次响应时长'], '排名': ''}],
        ignore_index=True)  # 计算全区
    tp1 = tp1.sort_values(by='首次响应时长', ascending=True).reset_index(drop=True)  # 按时长 排序
    tp1 = tp1.round({'首次响应时长': 2})  # 四舍五入
    table = pd.concat([table, table_temp[[col, '全流程及时率', '排名']], tp1], axis=1)
    return table


# 计算归档表
def make_excel(header, data, col, seven, off=1):
    table = pd.merge(header, data_total(data, col, '未归档总数', off=off), on=[col], how='left')
    # 缓装数
    table = pd.merge(table, data_total(data[data.定单状态 == '缓装'], col, '缓装数', off=off), on=[col], how='left')
    # 待装数
    table = pd.merge(table, data_total(data[data.定单状态 == '待装'], col, '待装数', off=off), on=[col], how='left')
    # 在途数
    table = pd.merge(table, data_total(data[data.定单状态 == '在途'], col, '在途数', off=off), on=[col], how='left')
    # 即办即装超时工单数
    table = pd.merge(table,
                     data_total(data[(data.BOSS派单时间 < str(chao8)) & (data.定单状态 == '在途')], col, '即办即装超时数', off=off),
                     on=[col], how='left')
    # 超8工作时工单数
    table = pd.merge(table, data_total(data[data.BOSS派单时间 < str(chao8work)], col, '超8工作时工单数', off=off), on=[col],
                     how='left')
    # 超72小时工单数
    table = pd.merge(table, data_total(data[data.BOSS派单时间 < str(chao72)], col, '超72小时工单数', off=off), on=[col],
                     how='left')
    # 超72小时工单数占比
    table['超72小时工单数占比'] = (table.超72小时工单数 / table.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    # 超7天工单数
    table = pd.merge(table, data_total(data[data.BOSS派单时间 < str(chao7day)], col, '超7天工单数', off=off), on=[col],
                     how='left')
    # 超30天工单数
    table = pd.merge(table, data_total(data[data.BOSS派单时间 < str(chao30day)], col, '超30天工单数', off=off), on=[col],
                     how='left')
    table = table.fillna(0)  # 批量替换nan
    # 昨20点后受理未预约
    table = pd.merge(table, data_total(
        data[(data.BOSS派单时间 > str(yesterday20)) & (data.定单状态 == '在途') & (data['第一次预约操作时间'].isnull())], col,
        '昨20点后受理未预约', off=off), on=[col], how='left')

    data1 = data[(data.定单状态 == '缓装') | (data.定单状态 == '待装')]

    data2 = data1[(data1.未完成大类 != '用户原因') & (data1.未完成大类 != '资源问题') & (data1.未完成大类 != '设备不足') & (data1.未完成大类 != '其他')]
    zt_total = pd.merge(header, data_total(data2, col, '在途数1', off=off), on=[col], how='left').fillna(0)

    data3 = data1[(data1.未完成大类 == '其他') & (data1['未完成分类'].str.contains('汛期') == False)]
    zt_total = pd.merge(zt_total, data_total(data3, col, '在途数2', off=off), on=[col], how='left').fillna(0)

    table['核减后在途数'] = table.在途数 + zt_total.在途数1 + zt_total.在途数2

    table = table.fillna(0)  # 批量替换nan
    table = pd.merge(table, data_total(seven, col, '七日均', off=off), on=[col], how='left')
    table['七日均'] = table.七日均 / 7
    table = table.round({'七日均': 0})  # 四舍五入
    table['压单比'] = table.核减后在途数 / table.七日均
    table['缓装工单占比'] = (table.缓装数 / table.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['待装工单占比'] = (table.待装数 / table.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['在途工单占比'] = (table.在途数 / table.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['未归档压单比'] = (table.未归档总数 / table.七日均).fillna(0)
    table = table.round({'压单比': 2, '未归档压单比': 2})  # 四舍五入
    table.drop(['超72小时工单数'], axis=1, inplace=True)
    table = table.fillna(0)  # 批量替换nan
    return table


# 退单分析
def analyse_excel(data, qt, ht, col, off=1):
    # 临时数据存放库
    storage_library = data_total(qt[qt.区域类型 == '城市'], col, '前台城镇退单数', off=off)
    storage_library = pd.merge(storage_library,
                               data_total(qt[qt.区域类型 == '农村'], col, '前台农村退单数', off=off), on=col,
                               how='left')
    storage_library = pd.merge(storage_library,
                               data_total(ht[ht.区域类型 == '城市'], col, '后台城镇退单数', off=off), on=col,
                               how='left')
    storage_library = pd.merge(storage_library,
                               data_total(ht[ht.区域类型 == '农村'], col, '后台农村退单数', off=off), on=col,
                               how='left')
    storage_library = pd.merge(storage_library, data_total(qt, col, '前台退单数', off=off), on=col, how='left')
    storage_library = pd.merge(storage_library, data_total(ht, col, '后台退单数', off=off), on=col, how='left')
    storage_library = storage_library.fillna(0)

    # 总竣工数
    table = data_total(data, col, '总竣工量', off=off)
    table['总装机数'] = table.总竣工量 + storage_library.前台退单数 + storage_library.后台退单数
    # 城镇竣工数
    table = pd.merge(table, data_total(data[data.区域类型 == '城镇'], col, '城镇竣工数', off=off),
                     on=col, how='left')
    # 农村竣工数
    table = pd.merge(table, data_total(data[data.区域类型 == '农村'], col, '农村竣工数', off=off),
                     on=col, how='left')
    # 退单数(城市)
    table['城市退单数'] = storage_library.前台城镇退单数 + storage_library.后台城镇退单数
    # 城市退单率
    table['城市退单率'] = (table.城市退单数 / (table.城镇竣工数 + table.城市退单数)).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 退单数(农村)
    table['农村退单数'] = storage_library.前台农村退单数 + storage_library.后台农村退单数
    # 农村退单率
    table['农村退单率'] = (table.农村退单数 / (table.农村竣工数 + table.农村退单数)).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 总退单率
    table['总退单率'] = ((table.城市退单数 + table.农村退单数) / table.总装机数).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 前台退单数，率
    table['前台退单数'] = storage_library['前台退单数']
    table['前台退单率'] = (table.前台退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    # 后台退单数，率
    table['后台退单数'] = storage_library['后台退单数']
    table['后台退单率'] = (table.后台退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    #
    qt.rename(columns={'前台退单二级分类': 'BOSS撤单原因'}, inplace=True)  # 改列名
    ht.rename(columns={'后台退单二级分类': '后端驳回原因'}, inplace=True)  # 改列名
    # 合并前后台退单详表
    all_table = ht.append(qt)
    # 建设原因退单数
    table = pd.merge(table, data_total(all_table[all_table.一级分类 == '建设原因'], col, '建设原因退单数', off=off),
                     on=col, how='left')
    # 其他原因退单数
    table = pd.merge(table, data_total(all_table[all_table.一级分类 == '其他原因'], col, '其他原因退单数', off=off),
                     on=col, how='left')
    # 前台原因退单数
    table = pd.merge(table, data_total(all_table[all_table.一级分类 == '前台原因'], col, '前台原因退单数', off=off),
                     on=col, how='left')
    # 网络原因退单数
    table = pd.merge(table, data_total(all_table[all_table.一级分类 == '网络原因'], col, '网络原因退单数', off=off),
                     on=col, how='left')
    # 用户原因退单数
    table = pd.merge(table, data_total(all_table[all_table.一级分类 == '用户原因'], col, '用户原因退单数', off=off),
                     on=col, how='left')
    table = table.fillna(0)
    # 退单率
    table['建设原因退单率'] = (table.建设原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['其他原因退单率'] = (table.其他原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['前台原因退单率'] = (table.前台原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['网络原因退单率'] = (table.网络原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['用户原因退单率'] = (table.用户原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    # 城市网络未覆盖退单数
    table = pd.merge(table,
                     data_total(all_table[(all_table.区域类型 == '城市') & (all_table.网络未覆盖 == '是')], col, '城市网络未覆盖退单数',
                                off=off), on=col, how='left')
    # 农村网络未覆盖退单数
    table = pd.merge(table,
                     data_total(all_table[(all_table.区域类型 == '农村') & (all_table.网络未覆盖 == '是')], col, '农村网络未覆盖退单数',
                                off=off), on=col, how='left')
    table = table.fillna(0)
    # 城市网络未覆盖退单率
    table['城市网络未覆盖退单率'] = (table['城市网络未覆盖退单数'] / (table['城镇竣工数'] + table['城市退单数'])).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 农村网络未覆盖退单率
    table['农村网络未覆盖退单率'] = (table['农村网络未覆盖退单数'] / (table['农村竣工数'] + table['农村退单数'])).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    return table


def checkbox_excel(data, qt, ht, col, table):
    # 临时数据存放库
    tem_p = checkbox_total(qt[qt.区域类型 == '城市'], col, '前台城镇退单数')
    table = table.append({'地市': '全区'}, ignore_index=True)

    storage_library = pd.merge(table, tem_p, on=['地市', col], how='left')

    tem_p = checkbox_total(qt[qt.区域类型 == '农村'], col, '前台农村退单数')
    storage_library = pd.merge(storage_library, tem_p, on=['地市', col], how='left')

    tem_p = checkbox_total(ht[ht.区域类型 == '城市'], col, '后台城镇退单数')
    storage_library = pd.merge(storage_library, tem_p, on=['地市', col], how='left')

    tem_p = checkbox_total(ht[ht.区域类型 == '农村'], col, '后台农村退单数')
    storage_library = pd.merge(storage_library, tem_p, on=['地市', col], how='left')

    tem_p = checkbox_total(qt, col, '前台退单数')
    storage_library = pd.merge(storage_library, tem_p, on=['地市', col], how='left')

    tem_p = checkbox_total(ht, col, '后台退单数')
    storage_library = pd.merge(storage_library, tem_p, on=['地市', col], how='left')
    storage_library = storage_library.fillna(0)
    # 总竣工数
    table = pd.merge(table, checkbox_total(data, col, '总竣工量'), on=['地市', col], how='left')
    # 总装机数
    table['总装机数'] = table.总竣工量 + storage_library.前台退单数 + storage_library.后台退单数
    # 城镇竣工数
    table = pd.merge(table, checkbox_total(data[data.区域类型 == '城镇'], col, '城镇竣工数'),
                     on=['地市', col], how='left')
    # 农村竣工数
    table = pd.merge(table, checkbox_total(data[data.区域类型 == '农村'], col, '农村竣工数'),
                     on=['地市', col], how='left')
    # 退单数(城市)
    table['城市退单数'] = storage_library.前台城镇退单数 + storage_library.后台城镇退单数
    # 城市退单率
    table['城市退单率'] = (table.城市退单数 / (table.城镇竣工数 + table.城市退单数)).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 退单数(农村)
    table['农村退单数'] = storage_library.前台农村退单数 + storage_library.后台农村退单数
    # 农村退单率
    table['农村退单率'] = (table.农村退单数 / (table.农村竣工数 + table.农村退单数)).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 总退单率
    table['总退单率'] = ((table.城市退单数 + table.农村退单数) / table.总装机数).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 前台退单数，率
    table['前台退单数'] = storage_library['前台退单数']
    table['前台退单率'] = (table.前台退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    # 后台退单数，率
    table['后台退单数'] = storage_library['后台退单数']
    table['后台退单率'] = (table.后台退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    #
    qt.rename(columns={'前台退单二级分类': 'BOSS撤单原因'}, inplace=True)  # 改列名
    ht.rename(columns={'后台退单二级分类': '后端驳回原因'}, inplace=True)  # 改列名
    # 合并前后台退单详表
    all_table = ht.append(qt)
    # 建设原因退单数
    table = pd.merge(table, checkbox_total(all_table[all_table.一级分类 == '建设原因'], col, '建设原因退单数'),
                     on=['地市', col], how='left')
    # 其他原因退单数
    table = pd.merge(table, checkbox_total(all_table[all_table.一级分类 == '其他原因'], col, '其他原因退单数'),
                     on=['地市', col], how='left')
    # 前台原因退单数
    table = pd.merge(table, checkbox_total(all_table[all_table.一级分类 == '前台原因'], col, '前台原因退单数'),
                     on=['地市', col], how='left')
    # 网络原因退单数
    table = pd.merge(table, checkbox_total(all_table[all_table.一级分类 == '网络原因'], col, '网络原因退单数'),
                     on=['地市', col], how='left')
    # 用户原因退单数
    table = pd.merge(table, checkbox_total(all_table[all_table.一级分类 == '用户原因'], col, '用户原因退单数'),
                     on=['地市', col], how='left')
    table = table.fillna(0)
    # 退单率
    table['建设原因退单率'] = (table.建设原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['其他原因退单率'] = (table.其他原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['前台原因退单率'] = (table.前台原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['网络原因退单率'] = (table.网络原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
    table['用户原因退单率'] = (table.用户原因退单数 / table.总装机数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))

    # 城市网络未覆盖退单数
    tem_p = checkbox_total(all_table[(all_table.区域类型 == '城市') & (all_table.网络未覆盖 == '是')], col, '城市网络未覆盖退单数')
    tem_p = tem_p.fillna(0)
    table = pd.merge(table, tem_p, on=['地市', col], how='left')
    tem_p = checkbox_total(all_table[(all_table.区域类型 == '农村') & (all_table.网络未覆盖 == '是')], col, '农村网络未覆盖退单数')
    tem_p = tem_p.fillna(0)
    table = pd.merge(table, tem_p, on=['地市', col], how='left')
    table = table.fillna(0)
    # 城市网络未覆盖退单率
    table['城市网络未覆盖退单率'] = (table['城市网络未覆盖退单数'] / (table['城镇竣工数'] + table['城市退单数'])).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    # 农村网络未覆盖退单率
    table['农村网络未覆盖退单率'] = (table['农村网络未覆盖退单数'] / (table['农村竣工数'] + table['农村退单数'])).fillna(0).apply(
        lambda x: '%.2f%%' % (x * 100))
    return table


city = pd.DataFrame({'地市': ['南宁', '桂林', '柳州', '玉林', '百色', '河池', '贵港', '钦州', '梧州', '北海', '崇左', '来宾', '贺州', '防城港', '全区']})

GROUP_PATH = f'{first_day1}-{yesterday1}家宽详表（集团模板）.xlsx'

GX_PATH = f'{first_day1}-{yesterday1}家宽详表（广西模板）千兆.xlsx'

CONTRAST_PATH = '程序专用对照表.xlsx'

ZT_PATH = f'{yesterday}宽带在途单统计.xls'

JG_PATH = f'{yesterday}宽带竣工单统计.xls'

QT_PATH = f'{first_day1}-{yesterday1}（前）退单一览表.xlsx'

HT_PATH = f'{first_day1}-{yesterday1}（后）退单一览表.xlsx'

WD_PATH = '工单图片智能质检结果报表-地市维度.xls'

# 整理表格数据

contrast_county = pd.read_excel(CONTRAST_PATH)  # 区县对照表
county = pd.DataFrame(contrast_county[['地市', '区县']])

davi = pd.DataFrame(contrast_county[['地市', '区县', '代维标段']])

contrast_grid = pd.read_excel(CONTRAST_PATH, sheet_name=1)  # 网格对照表

grid = pd.DataFrame(contrast_grid[['地市', '所属网格']])

print('正在读取宽带在途详表...')
zt_table = GetDataFrame(ZT_PATH)
list_ = list(zt_table.iloc[0])  # 取标段信息
zt_table.columns = list_  # 重命名 表头
zt_table = zt_table.drop(zt_table.head(1).index)  # 删除多余的 那一行标段信息
zt_table.dropna(subset=['工单编号'], inplace=True)  # 删除地市列为空的行
zt_table = zt_table.reset_index(drop=True)
# zt_table = pd.read_excel(ZT_PATH, skiprows=1)  # 在途详表
# zt_table.dropna(subset=['地市'], inplace=True)  # ,inplace=True  删除地市列为空的行
zt_table.loc[(zt_table['定单状态'] != '待装') & (zt_table['定单状态'] != '缓装'), '定单状态'] = '在途'
zt_table.rename(columns={'网格名称': '所属网格'}, inplace=True)  # 改列名
print('正在读取集团详表...')
group_table = GetDataFrame(GROUP_PATH)

# group_table = pd.read_excel(GROUP_PATH)  # 集团详表
group_table.rename(columns={'网格名称': '所属网格'}, inplace=True)  # 改列名
format_str(group_table, '工单历时减去总时长分化时', float)
format_str(group_table, '首次预约时长(H)', float)
group_table.loc[(group_table.区域类型 == '城镇') & (group_table.工单历时减去总时长分化时 <= 48), '是否超时'] = '未超时'
group_table.loc[(group_table.区域类型 == '农村') & (group_table.工单历时减去总时长分化时 <= 72), '是否超时'] = '未超时'
print('正在读取广西千兆详表...')
gx_1000 = GetDataFrame(GX_PATH)
# gx_1000 = pd.read_excel(GX_PATH)  # 广西千兆详表
gx_1000.rename(columns={'首次响应时长（H）': '首次响应时长'}, inplace=True)  # 改列名
gx_1000 = gx_1000[gx_1000.首次响应时长 >= 0]
format_str(gx_1000, '首次响应时长', float)
city_wd = pd.read_excel(WD_PATH, skiprows=2)  # 城市维度表
city_wd.rename(columns={'Unnamed: 0': '地市'}, inplace=True)  # 改列名
city_wd = city_wd[['地市', '分场景合格率']]
city_wd.loc[city_wd['地市'] == '总计', '地市'] = '全区'
# 修改网格信息
# 新匹配网格方法
# 在途
zt_table['区县'].fillna('区县信息为空', inplace=True)  # 批量替换nan
zt_table['所属网格'].fillna('网格信息为空', inplace=True)
map1 = pd.merge(zt_table, grid, on=["地市", '所属网格'])
map1.rename(columns={'工单编号': '工单编号1'}, inplace=True)  # 改列名
zt_table = pd.merge(zt_table, map1, how='left')
zt_table['工单编号1'].fillna('未匹配', inplace=True)  # 批量替换nan
zt_table.loc[(zt_table['工单编号1'] == '未匹配'), '所属网格'] = '网格信息为空'  # 网格进行判断，得到所筛选的
# 集团
group_table['区县'].fillna('区县信息为空', inplace=True)  # 批量替换nan
group_table['所属网格'].fillna('网格信息为空', inplace=True)
map1 = pd.merge(group_table, grid, on=["地市", '所属网格'])
map1.rename(columns={'boss工单号': 'boss工单号1'}, inplace=True)  # 改列名
group_table = pd.merge(group_table, map1, how='left')
group_table['boss工单号1'].fillna('未匹配', inplace=True)  # 批量替换nan
group_table.loc[(group_table['boss工单号1'] == '未匹配'), '所属网格'] = '网格信息为空'  # 网格进行判断，得到所筛选的

group_month = group_table[group_table.boss归档时间 > str(first_day)]  # 集团详表的当月数据
group_month = pd.merge(group_month, contrast_county, on=['地市', '区县'], how='left')  # 匹配代维信息
group_month.rename(columns={'代维公司_y': '代维公司', '首次预约时长(H)': '首次响应时长'}, inplace=True)
group_month.loc[(group_month['首次响应时长'] < 0) & (group_month['首次响应时长'] != -1), '首次响应时长'] = 0

group_month_drop = group_month[group_month.是否是校园宽带 == '否']  # 本月不含校园工单数据
print('正在读取宽带竣工详表...')
jg_table = pd.read_excel(JG_PATH, skiprows=1)
jg_table.dropna(subset=['地市'], inplace=True)  # ,inplace=True  删除地市列为空的行
jg_table.loc[jg_table['地市'] == '汇总', '地市'] = '全区'  # 修改地市列中的汇总为全区
jg_table = jg_table.fillna(0)  # 批量替换nan 为数字 0
print('正在读取前后台退单详表...')
# 退单详表
QT_chargeback = GetDataFrame(QT_PATH)
HT_chargeback = GetDataFrame(HT_PATH)


# 匹配网格信息
# 前台
QT_chargeback['区县'].fillna('区县信息为空', inplace=True)  # 批量替换nan
QT_chargeback['所属网格'].fillna('网格信息为空', inplace=True)
map1 = pd.merge(QT_chargeback, grid, on=["地市", '所属网格'])
map1.rename(columns={'BOSS工单号': 'BOSS工单号1'}, inplace=True)  # 改列名
QT_chargeback = pd.merge(QT_chargeback, map1, how='left')
QT_chargeback['BOSS工单号1'].fillna('未匹配', inplace=True)  # 批量替换nan
QT_chargeback['BOSS工单号1'] = QT_chargeback['BOSS工单号1'].astype(str)
QT_chargeback.loc[(QT_chargeback['BOSS工单号1'] == '未匹配'), '所属网格'] = '网格信息为空'  # 网格进行判断，得到所筛选的3
# 后台
HT_chargeback['区县'].fillna('区县信息为空', inplace=True)  # 批量替换nan
HT_chargeback['所属网格'].fillna('网格信息为空', inplace=True)
map1 = pd.merge(HT_chargeback, grid, on=["地市", '所属网格'])
map1.rename(columns={'BOSS工单号': 'BOSS工单号1'}, inplace=True)  # 改列名
HT_chargeback = pd.merge(HT_chargeback, map1, how='left')
HT_chargeback['BOSS工单号1'].fillna('未匹配', inplace=True)  # 批量替换nan
HT_chargeback['BOSS工单号1'] = HT_chargeback['BOSS工单号1'].astype(str)
HT_chargeback.loc[(HT_chargeback['BOSS工单号1'] == '未匹配'), '所属网格'] = '网格信息为空'  # 网格进行判断，得到所筛选的

# 前台退单原因
QT_reason = pd.read_excel(CONTRAST_PATH, sheet_name='前台退单原因分类')
QT_reason = QT_reason[['前台退单二级分类', '一级分类']]

# 后台退单原因
HT_reason = pd.read_excel(CONTRAST_PATH, sheet_name='后端退单原因分类')
HT_reason = HT_reason[['后台退单二级分类', '一级分类', '归类', '网络未覆盖']]

# 匹配前后台真实退单


QT_chargeback = format_str(QT_chargeback, '宽带号码')
HT_chargeback = format_str(HT_chargeback, '宽带号码')
group_month = format_str(group_month, '宽带号码')
group_month = format_str(group_month, '宽带号码')
group_month = format_str(group_month, '序号')
group_month = format_str(group_month, '序号')

QT_chargeback = pd.merge(QT_chargeback, group_month[['宽带号码', '序号']], on=['宽带号码'], how='left')
HT_chargeback = pd.merge(HT_chargeback, group_month[['宽带号码', '序号']], on=['宽带号码'], how='left')

QT_chargeback.序号_y.fillna('保留', inplace=True)  # 批量替换nan
HT_chargeback.序号_y.fillna('保留', inplace=True)

QT_chargeback = QT_chargeback[QT_chargeback.序号_y == '保留']
HT_chargeback = HT_chargeback[HT_chargeback.序号_y == '保留']

QT_chargeback.rename(columns={'BOSS撤单原因': '前台退单二级分类'}, inplace=True)  # 改列名
QT_chargeback = pd.merge(QT_chargeback, QT_reason, on=['前台退单二级分类'], how='left')
QT_chargeback['一级分类'].fillna('前台原因', inplace=True)

HT_chargeback.rename(columns={'后端驳回原因': '后台退单二级分类'}, inplace=True)  # 改列名
HT_chargeback = pd.merge(HT_chargeback, HT_reason[['后台退单二级分类', '一级分类', '网络未覆盖']], on=['后台退单二级分类'], how='left')

read_time = time.time()
print('读取并修改表格完毕!下面进行计算...读取表格耗时%0.0f秒钟ヾ(๑╹◡╹)ﾉ"' % (read_time - start_time))

############################################################################
# ######****************************************读取历史受理数据
LSshouli = pd.read_excel(f'{yesterday1}-月日报历史数据.xlsx', sheet_name='受理归档')
# ##用at函数取指定行指定列的元素
if LSshouli.at[1, '数据日期'] == datetime.datetime(before_yesterday.year, before_yesterday.month, before_yesterday.day, 23,
                                               59, 59):
    print(f'受理归档历史数据日期正确！可计算今日{today1}的受理归档平均数据')
else:
    print('受理归档历史数据日期错误！请检查！')
    f'{LSshouli}!!!@@@错误警告@@@ 受理归档历史数据日期错误！请检查！'

LSyadan = pd.read_excel(f'{yesterday1}-月日报历史数据.xlsx', sheet_name='压单比')
if (LSyadan.shape[0] - 1) == Y_day.day - 1:
    print(f'^_^月初到昨日{yesterday1}压单比数据已存在,可计算今日{today1}的月均压单比')
elif (LSyadan.shape[0] - 1) < Y_day.day - 1:
    print('压单比 数据缺失!!!')
    f'{LSyadan}!!!@@@错误警告@@@ 历史压单比 数据缺失!!!'
else:
    print('!!!@@@错误警告@@@ 历史压单比 数据多余于当天日报的需求!!!')
    f'{LSyadan}!!!@@@错误警告@@@ 历史压单比 数据多余于当天日报的需求!!!'
# #########*************************************************************************
LSBDyadan = pd.read_excel(f'{yesterday1}-月日报历史数据.xlsx', sheet_name='代维标段压单比')
if (LSyadan.shape[0] - 1) == Y_day.day - 1:
    print(f'^_^月初到昨日{yesterday1}代维标段压单比数据已存在,可计算今日{today1}的月均压单比')
elif (LSyadan.shape[0] - 1) < Y_day.day - 1:
    print('代维标段压单比 数据缺失!!!')
    f'{LSyadan}!!!@@@错误警告@@@ 历史代维标段压单比 数据缺失!!!'
else:
    print('!!!@@@错误警告@@@ 历史代维标段压单比 数据多余于当天日报的需求!!!')
    f'{LSyadan}!!!@@@错误警告@@@ 历史代维标段压单比 数据多余于当天日报的需求!!!'

LSDWyadan = pd.read_excel(f'{yesterday1}-月日报历史数据.xlsx', sheet_name='代维压单比')
if (LSyadan.shape[0] - 1) == Y_day.day - 1:
    print(f'^_^月初到昨日{yesterday1}代维公司压单比数据已存在,可计算今日{today1}的月均压单比')
elif (LSyadan.shape[0] - 1) < Y_day.day - 1:
    print('代维公司压单比 数据缺失!!!')
    f'{LSyadan}!!!@@@错误警告@@@ 历史代维公司压单比 数据缺失!!!'
else:
    print('!!!@@@错误警告@@@ 历史代维公司压单比 数据多余于当天日报的需求!!!')
    f'{LSyadan}!!!@@@错误警告@@@ 历史代维公司压单比 数据多余于当天日报的需求!!!'
# ##################### 第一页 ###########################
# 校园营销期间 用本句
daily_table1 = pd.DataFrame(columns=['地市', '当月日均办理', '当月日均归档', '宽带工单(不含校园办）', '校园工单办', '宽带工单(不含校园归）', '校园工单归'])
LSshouli_new = pd.merge(LSshouli[['地市', '当月办理', '当月归档', '当月日均办理', '当月日均归档', '数据日期']], jg_table, on=['地市'], how='left')
# 校园非营销期间 用本句
# daily_table1 = pd.DataFrame(columns=['地市', '当月日均办理', '当月日均归档', '受理工单(当日总量）', '网龄工单办', '归档工单(当日总量）', '网龄工单归'])
# LSshouli_new = LSshouli[['地市','当月办理','当月归档','当月日均办理','当月日均归档','数据日期']]
LSshouli_new['当月办理'] = pd.to_numeric(LSshouli_new.当月办理) + pd.to_numeric(LSshouli_new.工单受理量)
LSshouli_new['当月归档'] = pd.to_numeric(LSshouli_new.当月归档) + pd.to_numeric(LSshouli_new.工单竣工量)
LSshouli_new['当月日均办理'] = LSshouli_new.当月办理 / int(Y_day.strftime('%d'))
LSshouli_new['当月日均归档'] = LSshouli_new.当月归档 / int(Y_day.strftime('%d'))
LSshouli_new['数据日期'] = datetime.datetime(Y_day.year, Y_day.month, Y_day.day, 23, 59, 59)
daily_table1[['地市', '当月日均办理', '当月日均归档']] = LSshouli_new[['地市', '当月日均办理', '当月日均归档']]
# ###############################校园营销期间 用下面代码段
daily_table1['宽带工单(不含校园办）'] = pd.to_numeric(LSshouli_new.工单受理量) - pd.to_numeric(LSshouli_new.校园工单受理量)
daily_table1['校园工单办'] = pd.to_numeric(LSshouli_new.校园工单受理量)
daily_table1['宽带工单(不含校园归）'] = pd.to_numeric(LSshouli_new.工单竣工量) - pd.to_numeric(LSshouli_new.校园工单竣工量)
daily_table1['校园工单归'] = pd.to_numeric(LSshouli_new.校园工单竣工量)
# ##############################非校园营销期间 用本段代码
# daily_table1['受理工单(当日总量）'] = pd.to_numeric(LSshouli_new.工单受理量)
# daily_table1['网龄工单办'] = pd.to_numeric(LSshouli_new.包含网龄字段工单受理量)
# daily_table1['归档工单(当日总量）'] = pd.to_numeric(LSshouli_new.工单竣工量)
# daily_table1['网龄工单归'] = pd.to_numeric(LSshouli_new.包含网龄字段工单竣工量)

daily_table0 = daily_table1.round({'当月日均办理': 0, '当月日均归档': 0})  # 四舍五入
print('第一页受理及归档计算完毕')
# #########################################################第二页千兆宽带未归档工单总体情况 （不含校园工单）
zt_table_1000 = zt_table[(zt_table.是否校园宽带 == '否') & (zt_table.当前环节 != '外线施工（拆)')]  # 剔除校园工单数据
zt_table_1000 = zt_table_1000[zt_table_1000['定单主题'].str.contains('1000')]  # 筛选千兆数据
group_table_1000 = group_table[(group_table.boss归档时间 > str(seven_day)) & (group_table.是否是校园宽带 == '否')]
JG_seven = group_table_1000[group_table_1000['工单标题'].str.contains('1000')]  # 筛选千兆数据
daily_table1 = make_excel(city, zt_table_1000, '地市', JG_seven)
print('第二页千兆宽带未归档情况计算完毕')
# #########################################################第三页宽带未归档工单总体情况 （不含校园工单）
zt_table_drop = zt_table[(zt_table.是否校园宽带 == '否') & (zt_table.当前环节 != '外线施工（拆)')]  # 剔除校园工单数据
JG_seven = group_table[(group_table.boss归档时间 > str(seven_day)) & (group_table.是否是校园宽带 == '否')]  # 剔除校园工单数据
daily_table2 = make_excel(city, zt_table_drop, '地市', JG_seven)

JG_seven = group_table[(group_table.boss归档时间 > str(seven_day)) & (group_table['工单标题'].str.contains('新装'))]  # 剔除校园工单数据
JG_seven_day = JG_seven.groupby(['地市']).size().reset_index(name='新装七日均')
JG_seven_day = JG_seven_day.append([{'地市': '全区', '新装七日均': JG_seven_day.apply(lambda x: x.sum()).新装七日均}],
                                   ignore_index=True)

daily_table2 = pd.merge(daily_table2, JG_seven_day, on=['地市'], how='left')
daily_table2['新装七日均'] = daily_table2.新装七日均 / 7
daily_table2 = daily_table2.round({'新装七日均': 0})  # 四舍五入

print('第三页未归档总体情况计算完毕')
# #########################################################第四页未归档工单总体情况（校园工单）
zt_table_xy = zt_table[(zt_table.是否校园宽带 == '是') & (zt_table.当前环节 != '外线施工（拆)')]  # 校园工单数据
JG_seven = group_table[(group_table.boss归档时间 > str(seven_day)) & (group_table.是否是校园宽带 == '是')]  # 剔除校园工单数据
daily_table3 = make_excel(city, zt_table_xy, '地市', JG_seven)
print('第四页未归档总体情况计算完毕')
# #########################################################第五页每日压单比
Historical_Data = pd.read_excel(f'{yesterday1}-月日报历史数据.xlsx', sheet_name='压单比')

Historical_Data = Historical_Data.drop(Historical_Data.tail(1).index)  # 先删掉平均值那一行，从尾部开始算
Historical_Data.drop(['地市'], axis=1, inplace=True)
date = LSyadan.iloc[:, 0:1]  # 提取日期
date = date.drop(date.tail(1).index)  # 先删掉平均值那一行
Yadanbi = daily_table2[['地市', '未归档压单比']]
Yadanbi = pd.DataFrame(Yadanbi.values.T, index=Yadanbi.columns, columns=Yadanbi.index)  # 转置
list0 = list(Yadanbi.iloc[0])  # 取地市信息
Yadanbi.columns = list0  # 重命名 表头
Yadanbi = Yadanbi.drop(Yadanbi.head(1).index)  # 删除多余的 那一行地市信息
Yadanbi = Yadanbi.reset_index(drop=True)
temp = pd.concat([Historical_Data, Yadanbi], axis=0)  # 拼接 今日的压单比
temp.loc['平均'] = temp.apply(lambda x: x.mean())  # 列求均值
temp = temp.reset_index(drop=True)
date = date.append([{'地市': today1}], ignore_index=True)  # 添加日期
date = date.append([{'地市': '平均'}], ignore_index=True)  # 添加平均字段
daily_table4 = pd.concat([date, temp], axis=1)  # 拼接
print('第五页每日压单比计算完毕')
# #########################################################第六页未归档工单（城镇)(不含校园)
zt_table_city = zt_table_drop[zt_table_drop.区域属性 == '城镇']
JG_seven = group_table[
    (group_table.boss归档时间 > str(seven_day)) & (group_table.是否是校园宽带 == '否') & (group_table.区域类型 == '城镇')]
daily_table5 = make_excel(city, zt_table_city, '地市', JG_seven)
print('第六页未归档总体(城镇)计算完毕')

# #########################################################第七页未归档工单（农村）
zt_table_country = zt_table_drop[zt_table_drop.区域属性 == '农村']
JG_seven = group_table[
    (group_table.boss归档时间 > str(seven_day)) & (group_table.是否是校园宽带 == '否') & (group_table.区域类型 == '农村')]
daily_table6 = make_excel(city, zt_table_country, '地市', JG_seven)
print('第七页未归档总体(农村)计算完毕')
# #########################################################第八页家宽装机指标
daily_table7 = time_total(group_month_drop, '地市')
print('第八页家宽装机指标计算完毕')
# #########################################################第九页未归档工单(代维标段)
Biaoduan = pd.DataFrame({'代维标段': ['百色铁通', '北海铁通', '崇左铁通', '防城港铁通', '桂林怡创', '桂林铁通', '贵港浙邮', '贵港铁通', '河池铁通', '贺州铁通',
                                  '来宾铁通', '柳州铁通', '柳州宜通', '南宁铁通', '南宁润建', '钦州铁通', '梧州铁通', '玉林铁通', '代维信息为空(南宁)',
                                  '代维信息为空(桂林)', '代维信息为空(柳州)', '代维信息为空(玉林)', '代维信息为空(百色)', '代维信息为空(河池)', '代维信息为空(贵港)',
                                  '代维信息为空(钦州)', '代维信息为空(梧州)', '代维信息为空(北海)', '代维信息为空(崇左)', '代维信息为空(来宾)', '代维信息为空(贺州)',
                                  '代维信息为空(防城港)']})
zt_davi = pd.merge(zt_table, contrast_county, how='left')  # 匹配代维信息
zt_davi_drop = zt_davi[(zt_davi.是否校园宽带 == '否') & (zt_davi.当前环节 != '外线施工（拆)')]  # 剔除数据

group_davi = pd.merge(group_table, contrast_county, on=['地市', '区县'], how='left')  # 匹配代维信息
group_davi.rename(columns={'代维公司_y': '代维公司'}, inplace=True)

JG_seven = group_davi[(group_davi.boss归档时间 > str(seven_day)) & (group_davi.是否是校园宽带 == '否')]  # 剔除校园工单数据
daily_table8 = make_excel(Biaoduan, zt_davi_drop, '代维标段', JG_seven, off=0)
print('第九页未归档工单(代维标段)计算完毕')
# #########################################################第十页未归档工单(代维公司)
company = pd.DataFrame({'代维公司': ['润建', '铁通', '怡创', '宜通', '浙邮', '代维信息为空(南宁)', '代维信息为空(桂林)', '代维信息为空(柳州)', '代维信息为空(玉林)',
                                 '代维信息为空(百色)', '代维信息为空(河池)', '代维信息为空(贵港)', '代维信息为空(钦州)', '代维信息为空(梧州)', '代维信息为空(北海)',
                                 '代维信息为空(崇左)', '代维信息为空(来宾)', '代维信息为空(贺州)', '代维信息为空(防城港)']})
daily_table9 = make_excel(company, zt_davi_drop, '代维公司', JG_seven, off=0)
print('第十页未归档工单(代维公司)计算完毕')

# #########################################################第十一页压单比(代维标段)

section_new = total_ydb(LSBDyadan, daily_table8, '代维标段').tail(8)
daily_table10 = section_new  # 取最新几天的压单比
print('第十一页压单比(代维标段)计算完毕')
# #########################################################第十二页压单比(代维公司)

# .tail(8)
davi_new = total_ydb(LSDWyadan, daily_table9, '代维公司').tail(8)
daily_table11 = davi_new  # 取最新几天的压单比
print('第十二页压单比(代维公司)计算完毕')
# #########################################################第十三页代维标段装机指标
daily_table12 = time_total(group_month_drop, '代维标段')
print('第十三页代维标段装机指标计算完毕')
# #########################################################第十四页代维公司装机指标
daily_table13 = time_total(group_month_drop, '代维公司')
print('第十四页代维公司装机指标计算完毕')
# #########################################################第十五页未归档工单(区县)
daily_table14 = county  # 前两列 赋值为 地市 和 区县
# 未归档总数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop.groupby(['地市', '区县']).size().reset_index(name='未归档总数'), on=['地市', '区县'],
                         how='left')
# 缓装数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[zt_table_drop.定单状态 == '缓装'].groupby(['地市', '区县']).size().reset_index(name='缓装数'),
                         on=['地市', '区县'], how='left')
# 待装数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[zt_table_drop.定单状态 == '待装'].groupby(['地市', '区县']).size().reset_index(name='待装数'),
                         on=['地市', '区县'], how='left')
# 在途数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[zt_table_drop.定单状态 == '在途'].groupby(['地市', '区县']).size().reset_index(name='在途数'),
                         on=['地市', '区县'], how='left')
# 即办即装超时数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[(zt_table_drop.定单状态 == '在途') & (zt_table_drop.BOSS派单时间 < str(chao8))].groupby(
                             ['地市', '区县']).size().reset_index(name='即办即装超时数'), on=['地市', '区县'], how='left')
# 超8工作时工单数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao8work)].groupby(
                             ['地市', '区县']).size().reset_index(name='超8工作时工单数'), on=['地市', '区县'], how='left')
# 超72小时工单数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao72)].groupby(['地市', '区县']).size().reset_index(
                             name='超72小时工单数'), on=['地市', '区县'], how='left')
# 超72小时工单数占比
daily_table14['超72小时工单数占比'] = (daily_table14.超72小时工单数 / daily_table14.未归档总数).fillna(0).apply(
    lambda x: '%.2f%%' % (x * 100))

# 超7天工单数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao7day)].groupby(['地市', '区县']).size().reset_index(
                             name='超7天工单数'), on=['地市', '区县'], how='left')
# 超30天工单数
daily_table14 = pd.merge(daily_table14,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao30day)].groupby(
                             ['地市', '区县']).size().reset_index(
                             name='超30天工单数'), on=['地市', '区县'], how='left')
# 昨20点后受理未预约
daily_table14 = pd.merge(daily_table14, zt_table_drop[
    (zt_table_drop.BOSS派单时间 > str(yesterday20)) & (zt_table_drop.定单状态 == '在途') & (
        zt_table_drop['第一次预约操作时间'].isnull())].groupby(['地市', '区县']).size().reset_index(name='昨20点后受理未预约'),
                         on=['地市', '区县'], how='left')

# 在途数
data1 = zt_table_drop[(zt_table_drop.定单状态 == '缓装') | (zt_table_drop.定单状态 == '待装')]
data2 = data1[(data1.未完成大类 != '用户原因') & (data1.未完成大类 != '资源问题') & (data1.未完成大类 != '设备不足') & (data1.未完成大类 != '其他')]
data3 = data1[(data1.未完成大类 == '其他') & (data1['未完成分类'].str.contains('汛期') == False)]

zt_total = county
zt_total = pd.merge(zt_total, data2.groupby(['地市', '区县']).size().reset_index(name='在途数1'), on=['地市', '区县'], how='left')

zt_total = pd.merge(zt_total, data3.groupby(['地市', '区县']).size().reset_index(name='在途数2'), on=['地市', '区县'], how='left')

daily_table14 = pd.merge(daily_table14, zt_total, on=['地市', '区县'], how='left')

daily_table14 = daily_table14.fillna(0)  # 批量替换nan

daily_table14['核减后在途数'] = daily_table14.在途数 + daily_table14.在途数1 + daily_table14.在途数2

daily_table14.drop(['在途数1'], axis=1, inplace=True)

daily_table14.drop(['在途数2'], axis=1, inplace=True)

# 七日均
JG_seven_day = JG_seven.groupby(['地市', '区县']).size().reset_index(name='七日均')
daily_table14 = pd.merge(daily_table14, JG_seven_day, on=['地市', '区县'], how='left')

daily_table14['七日均'] = daily_table14.七日均 / 7
daily_table14 = daily_table14.round({'七日均': 0})  # 四舍五入

# 压单比
daily_table14['压单比'] = daily_table14.核减后在途数 / daily_table14.七日均

daily_table14['缓装工单占比'] = (daily_table14.缓装数 / daily_table14.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
daily_table14['待装工单占比'] = (daily_table14.待装数 / daily_table14.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
daily_table14['在途工单占比'] = (daily_table14.在途数 / daily_table14.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
daily_table14['未归档压单比'] = (daily_table14.未归档总数 / daily_table14.七日均).fillna(0)
daily_table14 = daily_table14.round({'压单比': 2, '未归档压单比': 2})  # 四舍五入
daily_table14.drop(['超72小时工单数'], axis=1, inplace=True)
daily_table14 = daily_table14.fillna(0)  # 批量替换nan
print('第十五页未归档工单(区县)计算完毕')
# #########################################################第十六页未归档工单(网格)
daily_table15 = grid  # 前两列 赋值为 地市 和 所属网格
# 未归档总数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop.groupby(['地市', '所属网格']).size().reset_index(name='未归档总数'), on=['地市', '所属网格'],
                         how='left')
# 缓装数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[zt_table_drop.定单状态 == '缓装'].groupby(['地市', '所属网格']).size().reset_index(
                             name='缓装数'),
                         on=['地市', '所属网格'], how='left')
# 待装数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[zt_table_drop.定单状态 == '待装'].groupby(['地市', '所属网格']).size().reset_index(
                             name='待装数'),
                         on=['地市', '所属网格'], how='left')
# 在途数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[zt_table_drop.定单状态 == '在途'].groupby(['地市', '所属网格']).size().reset_index(
                             name='在途数'),
                         on=['地市', '所属网格'], how='left')
# 即办即装超时数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[(zt_table_drop.定单状态 == '在途') & (zt_table_drop.BOSS派单时间 < str(chao8))].groupby(
                             ['地市', '所属网格']).size().reset_index(name='即办即装超时数'), on=['地市', '所属网格'], how='left')
# 超8工作时工单数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao8work)].groupby(
                             ['地市', '所属网格']).size().reset_index(name='超8工作时工单数'), on=['地市', '所属网格'], how='left')
# 超72小时工单数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao72)].groupby(['地市', '所属网格']).size().reset_index(
                             name='超72小时工单数'), on=['地市', '所属网格'], how='left')
# 超72小时工单数占比
daily_table15['超72小时工单数占比'] = (daily_table15.超72小时工单数 / daily_table15.未归档总数).fillna(0).apply(
    lambda x: '%.2f%%' % (x * 100))

# 超7天工单数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao7day)].groupby(
                             ['地市', '所属网格']).size().reset_index(
                             name='超7天工单数'), on=['地市', '所属网格'], how='left')
# 超30天工单数
daily_table15 = pd.merge(daily_table15,
                         zt_table_drop[zt_table_drop.BOSS派单时间 < str(chao30day)].groupby(
                             ['地市', '所属网格']).size().reset_index(
                             name='超30天工单数'), on=['地市', '所属网格'], how='left')
# 昨20点后受理未预约
daily_table15 = pd.merge(daily_table15, zt_table_drop[
    (zt_table_drop.BOSS派单时间 > str(yesterday20)) & (zt_table_drop.定单状态 == '在途') & (
        zt_table_drop['第一次预约操作时间'].isnull())].groupby(['地市', '所属网格']).size().reset_index(name='昨20点后受理未预约'),
                         on=['地市', '所属网格'], how='left')

# 在途数
zt_total1 = grid
zt_total1 = pd.merge(zt_total1, data2.groupby(['地市', '所属网格']).size().reset_index(name='在途数1'), on=['地市', '所属网格'],
                     how='left')

zt_total1 = pd.merge(zt_total1, data3.groupby(['地市', '所属网格']).size().reset_index(name='在途数2'), on=['地市', '所属网格'],
                     how='left')

daily_table15 = pd.merge(daily_table15, zt_total1, on=['地市', '所属网格'], how='left')

daily_table15 = daily_table15.fillna(0)  # 批量替换nan

daily_table15['核减后在途数'] = daily_table15.在途数 + daily_table15.在途数1 + daily_table15.在途数2

daily_table15.drop(['在途数1'], axis=1, inplace=True)
daily_table15.drop(['在途数2'], axis=1, inplace=True)

# 七日均
JG_seven_day = JG_seven.groupby(['地市', '所属网格']).size().reset_index(name='七日均')

daily_table15 = pd.merge(daily_table15, JG_seven_day, on=['地市', '所属网格'], how='left')
daily_table15['七日均'] = daily_table15.七日均 / 7
daily_table15 = daily_table15.round({'七日均': 0})  # 四舍五入
# 压单比
daily_table15['压单比'] = daily_table15.核减后在途数 / daily_table15.七日均

daily_table15['缓装工单占比'] = (daily_table15.缓装数 / daily_table15.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
daily_table15['待装工单占比'] = (daily_table15.待装数 / daily_table15.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
daily_table15['在途工单占比'] = (daily_table15.在途数 / daily_table15.未归档总数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
daily_table15['未归档压单比'] = (daily_table15.未归档总数 / daily_table15.七日均).fillna(0)
daily_table15 = daily_table15.round({'压单比': 2, '未归档压单比': 2})  # 四舍五入
daily_table15.drop(['超72小时工单数'], axis=1, inplace=True)
daily_table15 = daily_table15.fillna(0)  # 批量替换nan
print('第十六页未归档工单(所属网格)计算完毕')
# #########################################################第十七页区县装移指标

# 全流程时长
daily_table16 = group_month_drop[['地市', '区县', '代维标段', '工单历时减去总时长分化时']].groupby(
    ['地市', '区县', '代维标段']).mean().reset_index()
daily_table16.rename(columns={'工单历时减去总时长分化时': '全流程时长'}, inplace=True)  # 改列名
daily_table16 = daily_table16.round({'全流程时长': 2})  # 四舍五入
daily_table16 = pd.merge(davi, daily_table16, on=['地市', '区县', '代维标段'], how='left')

# 全流程及时率
temp1 = group_month_drop[['地市', '区县', '工单历时减去总时长分化时']].groupby(['地市', '区县']).size().reset_index(name='总工单数')
temp2 = group_month_drop[group_month_drop.是否超时 == '未超时'][['地市', '区县', '工单历时减去总时长分化时']].groupby(
    ['地市', '区县']).size().reset_index(
    name='未超时工单数')
daily_table16 = pd.merge(daily_table16, temp1, on=['地市', '区县'], how='left')
daily_table16 = pd.merge(daily_table16, temp2, on=['地市', '区县'], how='left')
daily_table16['全流程及时率'] = (daily_table16.未超时工单数 / daily_table16.总工单数).fillna(0).apply(
    lambda x: '%.2f%%' % (x * 100))  # 这里是计算各地市的及时率

# 首次响应时长
temp1 = group_month_drop[['地市', '区县', '首次响应时长']].groupby(['地市', '区县']).mean().reset_index()
daily_table16 = pd.merge(daily_table16, temp1, how='left', on=['地市', '区县'])
daily_table16 = daily_table16.round({'首次响应时长': 2})  # 四舍五入
daily_table16 = daily_table16.fillna(0)  # 批量替换nan
print('第十七页区县装移指标计算完毕')
# #########################################################第十八页网格装移指标
# 全流程时长
daily_table17 = group_month_drop[['地市', '所属网格', '工单历时减去总时长分化时']].groupby(
    ['地市', '所属网格']).mean().reset_index()
daily_table17.rename(columns={'工单历时减去总时长分化时': '全流程时长'}, inplace=True)  # 改列名
daily_table17 = daily_table17.round({'全流程时长': 2})  # 四舍五入
daily_table17 = pd.merge(grid, daily_table17, on=['地市', '所属网格'], how='left')
# 全流程及时率
temp3 = checkbox_total(group_month_drop, '所属网格', '总工单数')
temp4 = checkbox_total(group_month_drop[group_month_drop.是否超时 == '未超时'], '所属网格', '未超时工单数')
daily_table17 = pd.merge(daily_table17, temp3, on=['地市', '所属网格'], how='left')
daily_table17 = pd.merge(daily_table17, temp4, on=['地市', '所属网格'], how='left')
daily_table17['全流程及时率'] = (daily_table17.未超时工单数 / daily_table17.总工单数).fillna(0).apply(lambda x: '%.2f%%' % (x * 100))
# 首次响应时长
temp1 = group_month_drop[['地市', '所属网格', '首次响应时长']].groupby(['地市', '所属网格']).mean().reset_index()
daily_table17 = pd.merge(daily_table17, temp1, how='left', on=['地市', '所属网格'])
daily_table17 = daily_table17.round({'首次响应时长': 2})  # 四舍五入
daily_table17 = daily_table17.fillna(0)  # 批量替换nan
print('第十八页网格装移指标计算完毕')

# #########################################################第十九页地市退单分析
daily_table18 = analyse_excel(group_month, QT_chargeback, HT_chargeback, '地市')
daily_table18 = pd.merge(city, daily_table18, on='地市', how='left')
# 取在线信息
group_online = group_month[group_month['受理人员及联系电话'].str.contains('在线')]
group_online = pd.DataFrame(group_online)
group_online['在线'] = '在线'

QT_online = QT_chargeback[QT_chargeback['营业厅名称'].str.contains('在线')]
QT_online = pd.DataFrame(QT_online)
QT_online['在线'] = '在线'

HT_online = HT_chargeback[HT_chargeback['营业厅名称'].str.contains('在线')]
HT_online = pd.DataFrame(HT_online)
HT_online['在线'] = '在线'

daily_table_online = analyse_excel(group_online, QT_online, HT_online, '在线', off=0)
print('第十九页地市退单分析计算完毕')

# #########################################################第二十页区县退单分析
daily_table19 = checkbox_excel(group_month, QT_chargeback, HT_chargeback, '区县', county)
print('第二十页区县退单分析计算完毕')
# #########################################################第二十一页区县退单分析
daily_table20 = checkbox_excel(group_month, QT_chargeback, HT_chargeback, '所属网格', grid)
print('第二十一页网格退单分析计算完毕')
# #########################################################第二十二页装移指标(农村专项)
daily_table21 = time_total(group_month[group_month.区域类型 == '农村'], '地市')
print('第二十二页装移指标(农村专项)计算完毕')
# #########################################################第二十三页千兆退单
group_month_1000 = pd.DataFrame(group_month[group_month['工单标题'].str.contains('1000')])
QT_chargeback_1000 = pd.DataFrame(QT_chargeback[QT_chargeback['工单标题'].str.contains('1000')])
HT_chargeback_1000 = pd.DataFrame(HT_chargeback[HT_chargeback['工单标题'].str.contains('1000')])
daily_table22 = analyse_excel(group_month_1000, QT_chargeback_1000, HT_chargeback_1000, '地市')
daily_table22 = pd.merge(city, daily_table22, on='地市', how='left')
print('第二十三页千兆退单计算完毕')


# #########################################################通报
def rank_table(data, col, ascending=True):
    table = pd.DataFrame(data[['地市', col]])
    table['排名'] = table[col][:-1].rank(axis=0, ascending=ascending, method='dense')  # 输出排名
    table = table.sort_values(by='排名', ascending=False).reset_index(drop=True)  # 排序
    return table


# 把每个表格的数据分别取出排名排序
temp1 = rank_table(daily_table2, '压单比')  # 这里的temp只是一个临时变量

temp2 = pd.DataFrame(daily_table7.iloc[:, [6, 7]])
temp2 = pd.merge(city, temp2, on='地市', how='left')
temp2 = rank_table(temp2, '首次响应时长')

temp3 = pd.DataFrame(daily_table7.iloc[:, [0, 1]])
temp3 = pd.merge(city, temp3, on='地市', how='left')
temp3 = rank_table(temp3, '全流程时长')

temp4 = pd.DataFrame(daily_table7.iloc[:, [3, 4]])
temp4 = pd.merge(city, temp4, on='地市', how='left')

temp4['全流程及时率'] = temp4['全流程及时率'].str.strip("%").astype(float) / 100
temp4['排名'] = temp4['全流程及时率'][:-1].rank(axis=0, ascending=False, method='dense')  # 输出排名
temp4 = temp4.sort_values(by='排名', ascending=False).reset_index(drop=True)  # 排序
temp4['全流程及时率'] = temp4['全流程及时率'].apply(lambda x: '%.2f%%' % (x * 100))

temp5 = pd.DataFrame(daily_table18[['地市', '总退单率']])
temp5['总退单率'] = temp5['总退单率'].str.strip("%").astype(float) / 100
temp5['排名'] = temp5['总退单率'][:-1].rank(axis=0, ascending=True, method='dense')  # 输出排名
temp5 = temp5.sort_values(by='排名', ascending=False).reset_index(drop=True)  # 排序
temp5['总退单率'] = temp5['总退单率'].apply(lambda x: '%.2f%%' % (x * 100))

temp6 = rank_table(city_wd, '分场景合格率', False)

# 千兆首次响应时长
temp7 = gx_1000[['地市', '首次响应时长']].groupby(['地市']).mean().reset_index()
temp7['排名'] = temp7.rank(axis=0, ascending=True, method='dense').首次响应时长  # 输出排名
temp7.rename(columns={'首次响应时长': '千兆首次响应时长'}, inplace=True)  # 改列名
temp7 = temp7.sort_values(by='千兆首次响应时长', ascending=False).reset_index(drop=True)  # 按时长 排序
temp7 = temp7.append(
    [{'地市': '全区', '千兆首次响应时长': gx_1000[['地市', '首次响应时长']].mean().reset_index(name='千兆首次响应时长').at[0, '千兆首次响应时长'],
      '排名': ''}],
    ignore_index=True)  # 计算全区
temp7 = temp7.round({'千兆首次响应时长': 2})  # 四舍五入

daily_table23 = pd.concat([temp1, temp2, temp7, temp3, temp4, temp5, temp6], axis=1)

with pd.ExcelWriter('程序生成日报2021年' + today1 + '.xlsx') as writer:  # 写入结果为当前路径
    daily_table23.to_excel(writer, sheet_name='通报', startcol=0, startrow=1, index=False, header=True)
    daily_table0.to_excel(writer, sheet_name='1.受理及归档', startcol=0, index=False, header=True)
    daily_table1.to_excel(writer, sheet_name='2.千兆未归档情况', startcol=0, index=False, header=True)
    daily_table2.to_excel(writer, sheet_name='3.未归档总体情况', startcol=0, index=False, header=True)
    daily_table3.to_excel(writer, sheet_name='4.未归档总体情况(校园)', startcol=0, index=False, header=True)
    daily_table4.to_excel(writer, sheet_name='5.每日压单比', startcol=0, index=False, header=True)
    daily_table5.to_excel(writer, sheet_name='6.未归档城镇', startcol=0, index=False, header=True)
    daily_table6.to_excel(writer, sheet_name='7.未归档农村', startcol=0, index=False, header=True)
    daily_table7.to_excel(writer, sheet_name='8.家宽装移指标', startcol=0, index=False, header=True)
    daily_table8.to_excel(writer, sheet_name='9.代维标段未归档', startcol=0, index=False, header=True)
    daily_table9.to_excel(writer, sheet_name='10.代维公司未归档', startcol=0, index=False, header=True)
    daily_table10.to_excel(writer, sheet_name='11.代维标段压单比', startcol=0, index=False, header=True)
    daily_table11.to_excel(writer, sheet_name='12.代维公司压单比', startcol=0, index=False, header=True)
    daily_table12.to_excel(writer, sheet_name='13.代维标段装移指标', startcol=0, index=False, header=True)
    daily_table13.to_excel(writer, sheet_name='14.代维公司装移指标', startcol=0, index=False, header=True)
    daily_table14.to_excel(writer, sheet_name='15.区县未归档', startcol=0, index=False, header=True)
    daily_table15.to_excel(writer, sheet_name='16.网格未归档', startcol=0, index=False, header=True)
    daily_table16.to_excel(writer, sheet_name='17.区县装移指标', startcol=0, index=False, header=True)
    daily_table17.to_excel(writer, sheet_name='18.网格装移指标', startcol=0, index=False, header=True)
    daily_table18.to_excel(writer, sheet_name='19.退单分析', startcol=0, index=False, header=True)
    daily_table_online.to_excel(writer, sheet_name='19.退单分析', startcol=0, startrow=16, index=False, header=False)
    daily_table19.to_excel(writer, sheet_name='20.区县退单分析', startcol=0, index=False, header=True)
    daily_table20.to_excel(writer, sheet_name='21.网格退单分析', startcol=0, index=False, header=True)
    daily_table21.to_excel(writer, sheet_name='22.装移指标(农村专项)', startcol=0, index=False, header=True)
    daily_table22.to_excel(writer, sheet_name='23.千兆退单', startcol=0, index=False, header=True)

with pd.ExcelWriter(f'{today1}-月日报历史数据.xlsx') as writer:
    LSshouli_new.to_excel(writer, sheet_name='受理归档', startcol=0, index=False, header=True)
    daily_table4.to_excel(writer, sheet_name='压单比', startcol=0, index=False, header=True)
    section_new.to_excel(writer, sheet_name='代维标段压单比', startcol=0, index=False, header=True)
    davi_new.to_excel(writer, sheet_name='代维压单比', startcol=0, index=False, header=True)
print('表格计算已完毕,正在修改表格样式...')
print('ヾ(◍°∇°◍)ﾉﾞ')

path1 = ('程序生成日报2021年' + today1 + '.xlsx')
wb = load_workbook(path1)
sheet = wb.active

sheet.merge_cells('A1:U1')
sheet['A1'] = f'{yesterday}程序生成日报(通报)'

blue = PatternFill("solid", fgColor="C5D9F1")  # 浅蓝色
yellow = PatternFill("solid", fgColor="FFFF00")  # 黄色


def fillcolor(filename, color):
    for i, element1 in enumerate(filename):
        sheet.cell(row=element1[0], column=element1[1]).fill = color


deeporange = [(1, 1)]

fillcolor(deeporange, blue)


# 给表头字体加粗

def boldfont(row3, row4, column3, column4, size=10):
    for i in range(row3, row4):
        for j in range(column3, column4):
            sheet.cell(row=i, column=j).font = Font(name='微软雅黑', size=size, color="000000", b=True, i=False)


boldfont(1, 2, 1, 2, 16)

# 改变表格的字体
for i in range(2, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        sheet.cell(row=i, column=j).font = Font(name='微软雅黑', size=11, color="000000", b=False, i=False)
boldfont(2, 3, 1, sheet.max_column + 1, 10)
boldfont(17, 18, 1, sheet.max_column + 1, 10)

# 字体对齐方式
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
        sheet.cell(row=i, column=j).alignment = alignment

# 边框样式
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        side_border = Side(style="thin", color="000000")  # 黑色
        border = Border(left=side_border, right=side_border, top=side_border, bottom=side_border)
        sheet.cell(row=i, column=j).border = border

# 未归档压单比
for i in range(1, sheet.max_row - 2):
    ydb = sheet.cell(row=i + 2, column=2).value
    if ydb > 2:
        sheet.cell(row=i + 2, column=1).fill = yellow

# 首次响应时长
for i in range(1, sheet.max_row - 2):
    ydb = sheet.cell(row=i + 2, column=5).value
    if ydb > 2:
        sheet.cell(row=i + 2, column=4).fill = yellow

# 千兆首次响应时长
for i in range(1, sheet.max_row - 2):
    ydb = sheet.cell(row=i + 2, column=8).value
    if ydb > 0.5:
        sheet.cell(row=i + 2, column=7).fill = yellow

# 全流程时长
for i in range(1, sheet.max_row - 2):
    ydb = sheet.cell(row=i + 2, column=11).value
    if ydb > 48:
        sheet.cell(row=i + 2, column=10).fill = yellow

# 全流程及时率
for i in range(1, sheet.max_row - 2):
    ydb = sheet.cell(row=i + 2, column=14).value
    ydb = float(ydb.strip('%'))
    if ydb < 93:
        sheet.cell(row=i + 2, column=13).fill = yellow

# 总退单率
for i in range(1, sheet.max_row - 2):
    ydb = sheet.cell(row=i + 2, column=17).value
    ydb = float(ydb.strip('%'))
    if ydb > 20:
        sheet.cell(row=i + 2, column=16).fill = yellow

# 分场景合格率
for i in range(1, sheet.max_row - 2):
    ydb = sheet.cell(row=i + 2, column=20).value
    ydb = float(ydb.strip('%'))
    if ydb < 90:
        sheet.cell(row=i + 2, column=19).fill = yellow
wb.save(path1)
os.system('taskkill/IM et.exe /F')
quit()
end_time = time.time()
print('处理完毕!!!总耗时%0.0f秒钟ヾ(๑╹◡╹)ﾉ"' % (end_time - start_time))

