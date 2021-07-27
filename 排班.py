import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


def duibi(dataname):
    dataname = pd.DataFrame(dataname.values.T, index=dataname.columns, columns=dataname.index)  # 转置
    dataname = dataname.reset_index(drop=True)  # 此时索引为标段信息  重置索引
    dataname = pd.DataFrame(dataname.values.T, index=dataname.columns, columns=dataname.index)  # 转置
    return dataname


def color(filename, n):
    wb1 = openpyxl.load_workbook(filename)
    sheet1 = wb1['sheet1']
    fille = PatternFill("solid", fgColor="FFFF00")  # 红色
    for u, element1 in enumerate(n):
        sheet1.cell(row=element1[0], column=element1[1]).fill = fille
    wb1.save(filename)


# 旧排班
Schedule = pd.read_excel('员工排班表.xls', skiprows=3)
# 新排班
Schedule_New = pd.read_excel('排班表新模板-0518 - 副本.xlsx', sheet_name=1)
Schedule = Schedule.iloc[0:, 1:34]  # 切片

Schedule = Schedule[(Schedule.所属部门 == '审批组') | (Schedule.所属部门 == '催单组') | (Schedule.所属部门 == '一级调度')]

Schedule = Schedule.replace(
    {1: 'a', 2: 'b', 3: 'c', 4: 'd', 5: 'e', 6: 'f', 7: 'g', 8: 'h', 9: 'i', 10: 'j', 11: 'k', 25: '出差',
     26: '事1'})  # 替换数据

Schedule.fillna('休', inplace=True)  # 批量替换nan

Comparison = Schedule_New.iloc[0:, 10:41]
Comparison = Comparison.rename(columns=lambda x: 1)
for i, vl in enumerate(Comparison.columns.values):
    column_names = Comparison.columns.values
    column_names[i] = vl + i
    Comparison.columns = column_names

Name = Schedule_New[['姓名', '组别']]
Comparison = pd.concat([Name, Comparison], axis=1)
Comparison.rename(columns={'组别': '所属部门'}, inplace=True)  # 改列名
Comparison = Comparison.replace({'调休': '休', '季休': '休', '休（新）': '休'})

Estimate = pd.merge(Schedule, Comparison)

map1 = Comparison['姓名'].map(lambda x: x in list(Estimate['姓名']))  # 两个表之间一列进行判断
map1 = pd.DataFrame(map1)
map1.rename(columns={'姓名': '判断'}, inplace=True)  # 改列名
Comparison = pd.concat([Comparison, map1], axis=1)

Schedule = Schedule.reset_index(drop=True)  # 此时索引为标段信息  重置索引

list1 = list()
for i in Schedule.姓名:
    list1.append(i)

# ===============================================================================

Schedule = duibi(Schedule)
Comparison = duibi(Comparison)
a = []
b = []
for i in range(len(Schedule[0])):
    for j in range(len(Comparison[0])):
        if Schedule.iloc[i][0] == Comparison.iloc[j][0]:
            for k in range(33):
                if Schedule.iloc[i][k] != Comparison.iloc[j][k]:
                    # print(Schedule.iloc[i][0], Comparison.iloc[j][0])
                    # print(Schedule.iloc[i][k], Comparison.iloc[j][k])
                    # print(i, k)
                    a.append((i + 2, k + 1))  # 机器行列
                    b.append((j + 2, k + 1))  # 人工行列
                else:
                    continue
        else:
            continue

with pd.ExcelWriter('匹配后人工排班' + '.xlsx') as writer:
    Comparison.to_excel(writer, sheet_name='sheet1', startcol=0, index=False, header=True)  # 机器
with pd.ExcelWriter('匹配后机器排班' + '.xlsx') as writer:
    Schedule.to_excel(writer, sheet_name='sheet1', startcol=0, index=False, header=True)  # 人工

color("匹配后机器排班.xlsx", a)
color("匹配后人工排班.xlsx", b)
