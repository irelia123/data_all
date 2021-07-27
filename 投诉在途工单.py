# -*- coding: utf-8 -*-
"""
Created on Sun Apr 18 09:44:43 2021

@author: Administrator
"""

import os, glob
import time, datetime
import pandas as pd
import copy
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from openpyxl.styles import Font  # 字体样式
from openpyxl.styles import PatternFill  # 填充颜色
from openpyxl.styles import Alignment  # 字体对齐方式
from openpyxl.styles import Side, Border  # 边框样式
import warnings

warnings.filterwarnings('ignore')
print('输入一个数字N,表示N天前，如一天前输入1表示：today-1，N天前输入N表示：today-N')
numberDay = int(input('请在此输入N然后按回车键:'))  # 输入一个数字表示N天前
today = datetime.datetime.now()
yestday = today - datetime.timedelta(days=numberDay)  # 设置日期，用当天的日期减去x天得到预期的日期
Time = yestday.strftime('%Y-%m-%d')  # 格式为str
yestday0 = str(yestday.month) + '月' + yestday.strftime('%d') + '日'  # 日期转换str：9月01日


# Time= input('>>请输入年-月-日格式如：xxxx-xx-xx：')#输入时间格式例如：20201-04-27


def read_data(path, sheet):
    table = pd.read_excel(path, sheet_name=sheet, inplace=True)  # 读取excel表格
    return table


Gd_Zb = read_data('./' + Time + '归档.xlsx', 'Sheet0')  # 读取/eomsj表
Gd_Pb = read_data('./' + Time + '派发.xlsx', 'Sheet0')  # 读取eomsp表
Zt_Pb = read_data('./' + Time + '在途.xlsx', 'Sheet0')  # 读取eomszt表
Zj_Pb = read_data('./自定义数据.xlsx', '中间数据')  # 读取中间数据
Zdysj = read_data('./自定义数据.xlsx', '地市日报自定义数据')  # 读取自定义日报数据
Zd_Qx = read_data('./自定义数据.xlsx', '区县表自定义数据')  # 读取自定义区县数据
Zd_Wg = read_data('./自定义数据.xlsx', '网格表自定义数据')  # 读取自定网格义数据
Zd_Bxguishudi = read_data('./自定义数据.xlsx', '不详归属地市')  # 读取自定义不详归属地数据
Zd_jm = read_data('./自定义数据.xlsx', 'JM')  # 读取自定义不详归属地数据
guzhang = read_data('./自定义数据.xlsx', '故障不过夜月累计量')


# ---------------------------设置时间
def setTime(date, hour, minute, second):
    todayTime = datetime.datetime.strptime(date, '%Y-%m-%d')  # ---------------------手动输入日期
    todayTime = datetime.datetime(todayTime.year, todayTime.month, todayTime.day, hour, minute, second)  # 设置时间点
    return todayTime


todaytime = setTime(Time, 20, 00, 00)
Sz_Time = setTime(Time, 23, 59, 59)
Ju_Time = setTime('2020-08-31', 23, 59, 59)
Chaoshi2 = setTime(Time, 23, 59, 59)
zt_time = setTime(Time, 00, 00, 00)
todaytime1 = todaytime.strftime('%Y/%m/%d %X')  # 格式为str
Sz_Time1 = Sz_Time.strftime('%Y/%m/%d %X')  # 格式为str
zt_time1 = zt_time.strftime('%Y/%m/%d %X')  # 格式为str

print('时间设置完成')


# ---------剔除问题解决单$和家亲
def clear_data(retained_data):  # ---------剔除问题解决单$和家亲
    retained_data = retained_data.fillna('')
    retained_data['安装地址'] = ''
    retained_data['投诉内容'] = ''
    retained_data = retained_data[
        (retained_data.工单类型 != '问题解决单') & (retained_data.工单类型 != '和家亲')]  # ---------剔除问题解决单$和家亲
    retained_data = retained_data.reset_index(drop=True)  # 重新排序
    return retained_data


Gd_Zb = clear_data(Gd_Zb)
Gd_Pb = clear_data(Gd_Pb)
Zt_Pb = clear_data(Zt_Pb)
# 当日归档明细 
gdmx = Gd_Zb[['工单流水号']]
gdmx['是否超过二十点'] = ''
gdmx['是否超过二十四点'] = (((pd.to_datetime(todaytime) - pd.to_datetime(Gd_Zb.归档时间)).dt.days * 24 + (
            pd.to_datetime(todaytime) - pd.to_datetime(Gd_Zb.归档时间)).dt.seconds / 3600) < 0) & (((pd.to_datetime(
    Sz_Time) - pd.to_datetime(Gd_Zb.归档时间)).dt.days * 24 + (pd.to_datetime(Sz_Time) - pd.to_datetime(
    Gd_Zb.归档时间)).dt.seconds / 3600) > 0)
gdmx = gdmx.drop(['工单流水号'], axis=1)  # 删除列
Gd_Zb = pd.concat([gdmx, Gd_Zb], axis=1)  # 插入第一列为标段信息---------当日归档明细

# 当日派单明细

Sd = Gd_Pb[['首次到单时间']]
Sd['时间差'] = (pd.to_datetime(todaytime) - pd.to_datetime(Sd.首次到单时间)).dt.days * 24 + (
            pd.to_datetime(todaytime) - pd.to_datetime(
        Sd.首次到单时间)).dt.seconds / 3600  # 计算时间工公式    备注：公式是时间算的不能按元Excel数据修改
Sd['是否20点到达'] = Sd.时间差 > 0
Sd = Sd.drop(['首次到单时间', '时间差'], axis=1)  # 删除列

# 是否故障不过夜

Sd['是否故障不过夜'] = ((Gd_Pb['区域类型'] == '城市') & ((Gd_Pb['全球通身份等级'] == '全球通-钻石卡') | (Gd_Pb['速率'] == '1000M')))
# or ((Gd_Pb['区域类型'] =='城市')& (Gd_Pb['速率'] =='1000M'))
Gd_Pb = pd.concat([Sd, Gd_Pb], axis=1)  # 插入第一列为标段信息---------当日派单明细
# 是否故障不过夜


print('当日派单明细已完成')

# 当日在途明细

if list(Zt_Pb)[0] == '查询结果':  ###修改表头
    list0 = list(Zt_Pb.iloc[0])
    ###print('******请检查新的表头是否正确：',list0)
    Zt_Pb.columns = list0  ###重命名表头
    Zt_Pb.dropna(subset=['工单流水号'], inplace=True)  # ,inplace=True  删除空行
    Zt_Pb = Zt_Pb.drop(Zt_Pb[Zt_Pb.工单流水号 == '工单流水号'].index)  # 删除多余行

Zaitu_mx = Zt_Pb[['工单流水号']]

# 超过20:00:00

zaitu_cs = Zt_Pb[['工单流水号']]
zaitu_cs['是否超过二十点整'] = ((pd.to_datetime(todaytime) - pd.to_datetime(Zt_Pb.首次到单时间)).dt.days * 24 + (
            pd.to_datetime(todaytime) - pd.to_datetime(Zt_Pb.首次到单时间)).dt.seconds / 3600) > 0
zaitu_cs['是否超过二十四点整'] = (((pd.to_datetime(todaytime) - pd.to_datetime(Zt_Pb.首次到单时间)).dt.days * 24 + (
            pd.to_datetime(todaytime) - pd.to_datetime(Zt_Pb.首次到单时间)).dt.seconds / 3600) > 0) & (((pd.to_datetime(
    zt_time) - pd.to_datetime(Zt_Pb.首次到单时间)).dt.days * 24 + (pd.to_datetime(zt_time) - pd.to_datetime(
    Zt_Pb.首次到单时间)).dt.seconds / 3600) < 0)

zaitu_cs = zaitu_cs.drop(['工单流水号'], axis=1)  # 删除列

# ----------工单开始时间

Zaitu_Scddsj = Zt_Pb[['工单流水号', '首次到单时间']]
Zaitu_Scddsj.rename(columns={'首次到单时间': '工单开始时间'}, inplace=True)  # 改列名

# -------2021/4/27 24:00:00

Zaitu_sjc = Zt_Pb[['工单流水号']]
Zaitu_sjc['时间差'] = (pd.to_datetime(Sz_Time) - pd.to_datetime(Zt_Pb.首次到单时间)).dt.days * 24 + (
            pd.to_datetime(Sz_Time) - pd.to_datetime(
        Zt_Pb.首次到单时间)).dt.seconds / 3600  # 计算时间工公式    备注：公式是时间算的不能按元Excel数据修改

# --------------tag1

Zaitu_tag = Zaitu_sjc[['工单流水号', '时间差']]
Zaitu_tag['tag1'] = ''
Zaitu_tag['tag1'][Zaitu_tag['时间差'] >= 100] = '超100小时未竣工'  # 时间小于四天=超2-4日未竣工    备注：公式是时间算的不能按元Excel数据修改
Zaitu_tag['tag1'][
    (Zaitu_tag['时间差'] < 100) & (Zaitu_tag['时间差'] >= 48)] = '超48-100小时未竣工'  # 时间小于四天=超2-4日未竣工    备注：公式是时间算的不能按元Excel数据修改
Zaitu_tag['tag1'][
    (Zaitu_tag['时间差'] < 48) & (Zaitu_tag['时间差'] >= 24)] = '超24-48小时未竣工'  # 时间小于四天=超2-4日未竣工    备注：公式是时间算的不能按元Excel数据修改
Zaitu_tag = Zaitu_tag.drop(['时间差'], axis=1)  # 删除列

# ----------是否超24小时

Zaitu_Sfcessxs = Zaitu_sjc[['工单流水号', '时间差']]
Zaitu_Sfcessxs['是否超24小时'] = Zaitu_Sfcessxs['时间差'] > 24
Zaitu_Sfcessxs = Zaitu_Sfcessxs.drop(['时间差'], axis=1)  # 删除列

# -------------是否多次驳回

Zaitu_Sfdcbh = Zt_Pb[['工单流水号', '重派次数']]
Zaitu_Sfdcbh['是否多次驳回'] = (Zaitu_Sfdcbh['重派次数'] - 2) >= 0
Zaitu_Sfdcbh = Zaitu_Sfdcbh.drop(['重派次数'], axis=1)  # 删除列

# ------------区域类型

Zaitu_Qylx = Zt_Pb[['工单流水号', '区域类型']]

# ----------是否9月前工单
'''
Zaitu_Sfjyqgd = Zaitu_Scddsj[['工单流水号','工单开始时间']]
Zaitu_Sfjyqgd['是否9月前工单']  = (pd.to_datetime(Ju_Time)-pd.to_datetime(Zaitu_Sfjyqgd.工单开始时间)).dt.days*24+(pd.to_datetime(Ju_Time)-pd.to_datetime(Zaitu_Sfjyqgd.工单开始时间)).dt.seconds/3600     #计算时间工公式    备注：公式是时间算的不能按元Excel数据修改
Zaitu_Sfjyqgd['是否9月前工单'] = Zaitu_Sfjyqgd['是否9月前工单'] > 0
Zaitu_Sfjyqgd  = Zaitu_Sfjyqgd.drop(['工单开始时间'],axis = 1)#删除列'''

# 是否故障不过夜
zaitu_lastnight = Zt_Pb[['工单流水号']]
zaitu_lastnight['是否故障不过夜'] = ((Zt_Pb['区域类型'] == '城市') & ((Zt_Pb['全球通身份等级'] == '全球通-钻石卡') | (Zt_Pb['速率'] == '1000M')))
# -----------是否待质检

Zaitu_Sfdzj = Zt_Pb[['工单流水号', '工单状态']]
Zaitu_Sfdzj['是否待质检'] = Zaitu_Sfdzj['工单状态'] == '质检'
Zaitu_Sfdzj = Zaitu_Sfdzj.drop(['工单状态'], axis=1)  # 删除列

# ----------是否要剔除

Zaitu_Tichu1 = Zt_Pb[['工单流水号', '工单类型', '工单状态']]
Zaitu_Tichu1.ix[(Zaitu_Tichu1.工单类型 == ''), '是否要剔除'] = False
Zaitu_Tichu1.ix[(Zaitu_Tichu1.工单类型 == '家庭宽带'), '是否要剔除'] = False
Zaitu_Tichu1.ix[(Zaitu_Tichu1.工单状态 == '待归档'), '是否要剔除'] = True
Zaitu_Tichu1 = Zaitu_Tichu1.fillna(0)

Zaitu_Tichu = Zt_Pb[['工单流水号']]
zdy_jm = Zd_jm[['减免工单']]
zaitu_gdlshlist = list()
zdy_jmlist = list()
Zaitu_Tichulist = list()


def traversal(dataname, zaitulist):
    for i in dataname:  # 遍历dataFrame数据再放入列表中
        zaitulist.append(i)


traversal(Zaitu_Tichu.工单流水号, zaitu_gdlshlist)
traversal(zdy_jm.减免工单, zdy_jmlist)
traversal(Zaitu_Tichu1.是否要剔除, Zaitu_Tichulist)

for i in range(len(zaitu_gdlshlist)):
    for j in range(len(zdy_jmlist)):
        if zaitu_gdlshlist[i] == zdy_jmlist[j]:
            Zaitu_Tichulist[i] = True
            break
        else:
            if Zaitu_Tichulist[i] == 0:
                Zaitu_Tichulist[i] = False
                break

jm = pd.DataFrame(Zaitu_Tichulist, columns=['是否要剔除'])  # 把列表转化为dataFrame
Zaitu_Tichu['是否要剔除'] = jm[['是否要剔除']]

# -----------催单次数

Zaitu_Cdcs = Zt_Pb[['工单流水号', '催办次数', '客服催办次数']]
Zaitu_Cdcs['催单次数'] = Zaitu_Cdcs['催办次数'] + Zaitu_Cdcs['客服催办次数']
Zaitu_Cdcs = Zaitu_Cdcs.drop(['催办次数', '客服催办次数'], axis=1)  # 删除列

# ------------被客服催单

Zaitu_Bkfcd = Zt_Pb[['工单流水号', '客服催办次数']]

Zaitu_Bkfcd['被客服催单'] = (Zaitu_Bkfcd['客服催办次数'] - 1) >= 0
Zaitu_Bkfcd = Zaitu_Bkfcd.drop(['客服催办次数'], axis=1)  # 删除列

# ----------------是否超时

Zaitu_Sfcs = Zt_Pb[['工单流水号', '广西时限']]
Zaitu_Sfcs['是否超时'] = (pd.to_datetime(Chaoshi2) - pd.to_datetime(Zaitu_Sfcs.广西时限)).dt.days * 24 + (
            pd.to_datetime(Chaoshi2) - pd.to_datetime(
        Zaitu_Sfcs.广西时限)).dt.seconds / 3600  # 计算时间工公式    备注：公式是时间算的不能按元Excel数据修改
Zaitu_Sfcs['是否超时'] = Zaitu_Sfcs['是否超时'] + 8
Zaitu_Sfcs = Zaitu_Sfcs.drop(['广西时限'], axis=1)  # 删除列

# 20:00

zaitu_Esdz = Zt_Pb[['工单流水号']]
Zaitu_Bdll = Zt_Pb[['责任代维班组']]
Zhongjian_Kl = Zj_Pb[['代维班组', '地市代维一']]
list1 = list()  # 创建三个空列表
list2 = list()
list3 = list()
for i in Zaitu_Bdll.责任代维班组:  # 遍历dataFrame数据再放入列表中
    list1.append(i)
for j in Zhongjian_Kl.代维班组:
    list2.append(j)
for k in Zhongjian_Kl.地市代维一:
    list3.append(k)

list4 = list()
for i in range(len(list1)):
    for j in range(len(list2)):
        if list1[i] == list2[j]:
            list4.append(list3[j])
            break
    else:
        list4.append('')
df = pd.DataFrame(list4, columns=['责任代维班组'])  # 把列表转化为dataFrame
zaitu_Esdz['20:00'] = df[['责任代维班组']]

# ------------地市

zaitu_Dishi = Zt_Pb[['工单流水号']]
list5 = list()
list6 = list()
list7 = list()
list8 = list()
list9 = list()
list10 = list()
list11 = list()
zhongjian_Lm = Zj_Pb[['地市代维一', '地市二']]
for i in zhongjian_Lm.地市代维一:
    list5.append(i)
for i in zhongjian_Lm.地市二:
    list6.append(i)

for i in Zt_Pb.责任地市:
    list8.append(i)
for i in Zd_Bxguishudi.工单流水号:
    list9.append(i)
for i in Zd_Bxguishudi.责任地市:
    list10.append(i)
for i in Zt_Pb.工单流水号:
    list11.append(i)

for i in range(len(list4)):
    for j in range(len(list5)):
        if list4[i] == list5[j]:
            list7.append(list6[j])
            break
        else:
            continue
    else:
        list7.append(0)
for i in range(len(list4)):
    if list7[i] == 0:
        if list8[i] != '':
            list7[i] = list8[i]
        else:
            continue
    else:
        continue
for i in range(len(list4)):
    for h in range(len(list9)):
        if list7[i] == 0:
            if list11[i] == list9[h]:
                list7[i] = list10[h]
                break
            else:
                continue
        else:
            break

cf = pd.DataFrame(list7, columns=['地市'])  # 把列表转化为dataFrame
zaitu_Dishi['地市'] = cf[['地市']]
print('地市已完成')

# ------------8:00

zaitu_Bdz = Zt_Pb[['工单流水号']]
list_g = list()
list_h = list()
list_l = list()
list_n = list()
list_i = list()
list_j = list()
list_Bo = list()
list_Br = list()
list_Bdz = list()
zhongjian_Ln = Zj_Pb[['维护区域二', '代维一', '地市代维一', '简称一', '代维公司', '简称']]
for i in zhongjian_Ln.维护区域二:
    list_g.append(i)
for i in zhongjian_Ln.代维一:
    list_h.append(i)
for i in zhongjian_Ln.地市代维一:
    list_l.append(i)
for i in zhongjian_Ln.简称一:
    list_n.append(i)
for i in zhongjian_Ln.代维公司:
    list_i.append(i)
for i in zhongjian_Ln.简称:
    list_j.append(i)

for i in Zt_Pb.责任区县:
    list_Bo.append(i)
for i in Zt_Pb.客服最后重派时间:
    list_Br.append(i)

for i in range(len(list_Bo)):
    for j in range(len(list_g)):
        if list_Bo[i] == list_g[j]:
            list_Bdz.append(list_h[j])
            break
        else:
            continue
    else:
        list_Bdz.append('')
for i in range(len(list4)):
    for j in range(len(list_l)):
        if list_Bdz[i] == '':
            if list4[i] == list_l[j]:
                list_Bdz[i] = list_n[j]
                break
            else:
                continue
        else:
            break
for i in range(len(list_Br)):
    for j in range(len(list_i)):
        if list_Bdz[i] == '':
            if list_Br[i] == list_i[j]:
                list_Bdz[i] = list_j[j]
                break
            else:
                continue
        else:
            break

bdz = pd.DataFrame(list_Bdz, columns=['八点整'])  # 把列表转化为dataFrame
zaitu_Bdz['8:00'] = bdz[['八点整']]
print('8:00数据处理完成')

# ----------代维最终

zaitu_Dwzz = Zt_Pb[['工单流水号']]
list_Dwzz = list()
list_p = list()
list_q = list()
for i in Zj_Pb.地市三:
    list_p.append(i)
for i in Zj_Pb.代维二:
    list_q.append(i)
for i in range(len(list4)):
    for j in range(len(list_l)):
        if list4[i] == list_l[j]:
            list_Dwzz.append(list_n[j])
            break
        else:
            continue
    else:
        list_Dwzz.append('')
for i in range(len(list7)):
    for j in range(len(list_p)):
        if list_Dwzz[i] == '':
            if list7[i] == list_p[j]:
                list_Dwzz[i] = list_q[j]
                break
            else:
                continue
        else:
            break
for i in range(len(list_Bdz)):
    if list_Dwzz[i] == '':
        list_Dwzz[i] = list_Bdz[i]
    else:
        continue
for i in range(len(list_Dwzz)):
    if list_Dwzz[i] == '':
        list_Dwzz[i] = '铁通'
        continue
    else:
        continue

dwzz = pd.DataFrame(list_Dwzz, columns=['代维最终'])  # 把列表转化为dataFrame
zaitu_Dwzz['代维(最终)'] = dwzz[['代维最终']]

# ------------地市代维最终

zaitu_Dsdwzz = Zt_Pb[['工单流水号']]
list_Dsdwzz = list()
for i in range(len(list7)):
    if list_Dwzz[i] == list_Bdz[i]:
        if list4[i] != '':
            list_Dsdwzz.append(list4[i])
        else:
            list_Dsdwzz.append(list7[i] + list_Dwzz[i])
    else:
        list_Dsdwzz.append(list7[i] + list_Dwzz[i])
dsdwzz = pd.DataFrame(list_Dsdwzz, columns=['地市代维最终'])  # 把列表转化为dataFrame
zaitu_Dsdwzz['地市代维(最终)'] = dsdwzz[['地市代维最终']]

# -------质检时限

zaitu_Zjsx = Zt_Pb[['工单流水号']]
gxsx = Zt_Pb[['广西时限']]
list_Zjsx = list()
for i in gxsx.广西时限:
    strTime = i
    startTime1 = datetime.datetime.strptime(strTime, "%Y-%m-%d %X")  # 把strTime转化为时间格式,后面的秒位自动补位的
    if '20:00:00' in i:
        startTime2 = (startTime1 + datetime.timedelta(hours=16)).strftime("%Y-%m-%d %X")  # 把startTime时间加16小时
        list_Zjsx.append(startTime2)
    else:
        startTime2 = (startTime1 + datetime.timedelta(hours=4)).strftime("%Y-%m-%d %X")  # 把startTime时间加4小时
        list_Zjsx.append(startTime2)
zjsx = pd.DataFrame(list_Zjsx, columns=['质检时限'])
zaitu_Zjsx['质检时限'] = zjsx[['质检时限']]


# ----------------拼接当日在途明细

def joint(name1, name2, str1):  # 拼接函数
    name1 = pd.merge(name1, name2, on=[str1], how='left')
    return name1


list_zaitu = [Zaitu_mx, zaitu_Dishi, Zaitu_Sfcessxs, Zaitu_Sfdcbh, Zaitu_Scddsj, Zaitu_Qylx, zaitu_Dsdwzz, zaitu_Dwzz,
              zaitu_Bdz, zaitu_Esdz, Zaitu_sjc, Zaitu_tag, zaitu_Zjsx, zaitu_lastnight, Zaitu_Sfdzj, Zaitu_Tichu,
              Zaitu_Cdcs, Zaitu_Bkfcd, Zaitu_Sfcs, Zt_Pb]
for i in range(1, (len(list_zaitu))):
    list_zaitu[0] = joint(list_zaitu[0], list_zaitu[i], '工单流水号')
Zaitu_mx = list_zaitu[0]
Drztmx = Zaitu_mx.columns.tolist()  # 把Cols1的列名称，取出来放到一个list里边。即返回['a', 'b', 'c', 'd', 'e', '责任地市']
Drztmx.insert(19, Drztmx.pop(Drztmx.index('工单流水号')))  # pop()把工单开始时间从cols列表里挖出来，通过位置参数“0”，然后放到第一列。
Zaitu_mx = Zaitu_mx[Drztmx]  # 排列后数据。
Gd_Sj = copy.deepcopy(Zaitu_mx)
Gd_Sj = pd.concat([zaitu_cs, Gd_Sj], axis=1)  # 插入第一列为标段信息---------当日派单明细

# -----------------------------------------------地市日报

Dishi = pd.DataFrame(
    {'地市': ['南宁', '桂林', '柳州', '玉林', '百色', '河池', '贵港', '钦州', '梧州', '北海', '崇左', '来宾', '贺州', '防城港', '全区']},
    pd.Index(range(15)))

# 当日归档

Temp1 = Gd_Zb.groupby(['责任地市']).size().reset_index(name='当日归档')
Temp1 = Temp1.append([{'责任地市': '全区', '当日归档': Temp1.apply(lambda x: x.sum()).当日归档}], ignore_index=True)
Temp1.rename(columns={'责任地市': '地市'}, inplace=True)  # 改列名
Temp1 = pd.merge(Dishi, Temp1, on=['地市'], how='left')  # 拼接
print('当日归档已完成')

# 当日派单

Temp2 = Gd_Pb.groupby(['责任地市']).size().reset_index(name='当日派单')
Temp2 = Temp2.append([{'责任地市': '全区', '当日派单': Temp2.apply(lambda x: x.sum()).当日派单}], ignore_index=True)
Temp2.rename(columns={'责任地市': '地市'}, inplace=True)  # 改列名
print('当日派单已完成')

# 故障不过夜20点前派单量

esdqpdl = Gd_Pb[(Gd_Pb.是否20点到达 == True) & (Gd_Pb.是否故障不过夜 == True)].groupby(['责任地市']).size().reset_index(name='二十点前派单量')
esdqpdl = esdqpdl.append([{'责任地市': '全区', '二十点前派单量': esdqpdl.apply(lambda x: x.sum()).二十点前派单量}], ignore_index=True)
esdqpdl.rename(columns={'责任地市': '地市'}, inplace=True)  # 改列名
esdqpdl = pd.merge(Dishi, esdqpdl, on=['地市'], how='left')  # 拼接
esdqpdl.fillna(0, inplace=True)

# 故障不过夜在途

rbzaitu = Gd_Sj[(Gd_Sj.是否超过二十四点整 == True) & (Gd_Sj.是否故障不过夜 == True) & (Gd_Sj.是否要剔除 == False)].groupby(
    ['地市']).size().reset_index(name='故障不过夜在途')
rbzaitu = rbzaitu.append([{'地市': '全区', '故障不过夜在途': rbzaitu.apply(lambda x: x.sum()).故障不过夜在途}], ignore_index=True)
rbzaitu = pd.merge(Dishi, rbzaitu, on=['地市'], how='left')  # 拼接
rbzaitu.fillna(0, inplace=True)

# 当日20~24点归档
rbdrgd = Gd_Zb[Gd_Zb.是否超过二十四点 == True].groupby(['责任地市']).size().reset_index(name='当日二十点到二十四点归档')
rbdrgd.rename(columns={'责任地市': '地市'}, inplace=True)  # 改列名
rbdrgd = rbdrgd.append([{'地市': '全区', '当日二十点到二十四点归档': rbdrgd.apply(lambda x: x.sum()).当日二十点到二十四点归档}], ignore_index=True)
rbdrgd = pd.merge(Dishi, rbdrgd, on=['地市'], how='left')  # 拼接
rbdrgd.fillna(0, inplace=True)
# 故障不过夜月累计

guzhang = guzhang.fillna(0)
gzpd = guzhang.iloc[0:16, 0:]
gzzt = guzhang.iloc[18:34, 0:]
gzjsl = guzhang.iloc[36:52, 0:]
gzldata = guzhang.iloc[0:1, 1:-2]
gzldata = pd.DataFrame(gzldata.values.T, index=gzldata.columns, columns=gzldata.index)  # 转置
gzllist = list()
for i in gzldata[0]:
    gzllist.append(i)


# 重命名表头

def ResetIndex(gzlist):
    gzlist = gzlist.reset_index(drop=True)
    list0 = list(gzlist.iloc[0])  # 取地市信息
    gzlist.columns = list0  # 重命名 表头
    gzlist = gzlist.drop(gzlist.head(1).index)  # 删除多余的 那一行地市信息
    gzlist = gzlist.reset_index(drop=True)
    return gzlist


guzhangpd = ResetIndex(gzpd)
guzhangzt = ResetIndex(gzzt)
guzhangjsl = ResetIndex(gzjsl)
guzhangpd0 = copy.deepcopy(guzhangpd)
guzhangzt0 = copy.deepcopy(guzhangzt)
guzhangjsl0 = copy.deepcopy(guzhangjsl)
guzhangpd0 = guzhangpd0.drop(['合计', '地市'], axis=1)  # 删除列
guzhangzt0 = guzhangzt0.drop(['合计', '地市'], axis=1)  # 删除列
guzhangjsl0 = guzhangjsl0.drop(['合计', '地市'], axis=1)  # 删除列
guzhangpd0[yestday0] = esdqpdl[['二十点前派单量']]
guzhangzt0[yestday0] = rbzaitu[['故障不过夜在途']]
guzhangpd0.fillna(0, inplace=True)
guzhangzt0.fillna(0, inplace=True)

guzhangjsl0[yestday0] = ((guzhangpd0[yestday0] - guzhangzt0[yestday0]) / guzhangpd0[yestday0]).apply(
    lambda x: '%.2f%%' % (x * 100))

guzhangpd0['合计'] = guzhangpd['合计'] + esdqpdl['二十点前派单量']
guzhangzt0['合计'] = guzhangzt['合计'] + rbzaitu['故障不过夜在途']
guzhangjsl0['合计'] = ((guzhangpd0['合计'] - guzhangzt0['合计']) / guzhangpd0['合计']).apply(lambda x: '%.2f%%' % (x * 100))

guzhangpd0 = pd.concat([Dishi, guzhangpd0], axis=1)  # 插入第一列为地市信息
guzhangzt0 = pd.concat([Dishi, guzhangzt0], axis=1)  # 插入第一列为地市信息
guzhangjsl0 = pd.concat([Dishi, guzhangjsl0], axis=1)  # 插入第一列为地市信息
guzhangpd0 = pd.concat([guzhangpd0, Dishi], axis=1)  # 插入第一列为地市信息
guzhangzt0 = pd.concat([guzhangzt0, Dishi], axis=1)  # 插入第一列为地市信息
guzhangjsl0 = pd.concat([guzhangjsl0, Dishi], axis=1)  # 插入第一列为地市信息
for i in gzllist:
    guzhangjsl0[i] = guzhangjsl0[i].apply(lambda x: '%.2f%%' % (x * 100))

print('故障不过夜月累计完成')

# 处理及时率-月累计

rbchulijsl = Dishi[['地市']]
rbchulijsl['故障不过夜处理及时率'] = guzhangjsl0[['合计']]

# 导入自定义数据

Zdysj.分母 = Zdysj.分母.fillna('')
Zdysj.归档总量 = Zdysj.归档总量.fillna('')
Zdysj.日 = Zdysj.日.fillna('')
Zdysj.时 = Zdysj.时.fillna('')

# 24点在途

Temp5 = Gd_Sj[Gd_Sj.是否要剔除 == False].groupby(['地市']).size().reset_index(name='二十四点在途')
Temp5 = Temp5.append([{'地市': '全区', '二十四点在途': Temp5.apply(lambda x: x.sum()).二十四点在途}], ignore_index=True)
Temp5 = pd.merge(Dishi, Temp5, on=['地市'], how='left')  # 拼接
print('24点在途已完成')

# 压单比<=0.8

Temp5['压单比≤0.8'] = Temp5.二十四点在途 / Zdysj.日均归档量
Temp5['压单比≤0.8'] = Temp5['压单比≤0.8'].round(decimals=2)  # 保留两位小数
print('压单比≤0.8已完成')

# 在途目标

Temp6 = Dishi[['地市']]
Temp6['日均归档量'] = Zdysj[['日均归档量']]
Temp6['在途目标'] = Temp6.日均归档量 * 0.8
Temp6.在途目标 = Temp6.在途目标 // 1  # 保留0位小数
Temp6 = Temp6.drop(['日均归档量'], axis=1)  # 删除列
print('在途目标已完成')


# -------------超长超时工单

def overtime(time1, deadline):
    Temp = Gd_Sj[(Gd_Sj.是否要剔除 == False) & (Gd_Sj.时间差 > time1)].groupby(['地市']).size().reset_index(name='超过N小时')
    Temp = Temp.append([{'地市': '全区', '超过N小时': Temp.apply(lambda x: x.sum()).超过N小时}], ignore_index=True)
    Temp.超过N小时 = Temp.超过N小时.fillna(0)
    Temp = pd.merge(Dishi, Temp, on=['地市'], how='left')  # 拼接
    Temp.rename(columns={'超过N小时': deadline}, inplace=True)  # 改列名
    return Temp


Temp8 = overtime(48, '超过四十八小时')
Temp9 = overtime(168, '超过七日')
Temp10 = overtime(360, '超过十五日')

# 超48小时占比

Temp13 = Dishi[['地市']]
Temp13['超四十八小时占比'] = (Temp8.超过四十八小时 / Temp5.二十四点在途).fillna(0)

Temp13['超四十八小时占比'] = Temp13['超四十八小时占比'].apply(lambda x: '%.f%%' % (x * 100))

print('超48小时占比已完成')

# -----------------------------------拼接地市日报

Zdysj['全流程处理时长≤16'] = Zdysj['全流程处理时长≤16'].round(decimals=2)  # 保留两位小数
Zdysj['30分钟投诉首响及时率≥95%'] = Zdysj['30分钟投诉首响及时率≥95%'].apply(lambda x: '%.2f%%' % (x * 100))
Zdysj['全流程处理及时率≥93%'] = Zdysj['全流程处理及时率≥93%'].apply(lambda x: '%.2f%%' % (x * 100))
Zdysj['日均归档量'] = Zdysj['日均归档量'].round(decimals=2)  # 保留两位小数

Dishi1 = copy.deepcopy(Dishi)
Dishi1['地市1'] = Dishi[['地市']]
list_cityrb = [Dishi, esdqpdl, rbzaitu, rbdrgd, Temp2, Temp1, Temp5, Temp6, Dishi1, Temp8, Temp9, Temp10, Temp13,
               Dishi1, rbchulijsl]
for i in range(1, len(list_cityrb)):
    list_cityrb[0] = joint(list_cityrb[0], list_cityrb[i], '地市')
Dishi_Rb = copy.deepcopy(list_cityrb[0])
zd_ysj = copy.deepcopy(Zdysj)
zd_ysj = zd_ysj.drop(['昨日在途'], axis=1)  # 删除列
Dishi_Rb = pd.concat([Dishi_Rb, zd_ysj], axis=1)  # 插入第一列为标段信息---------当日派单明细
Dishi_Rb = Dishi_Rb.fillna(0)
Dishi_Rb.rename(columns={'二十点前派单量': '故障不过夜20点前派单量', '地市_x': '地市', '地市_y': '地市'}, inplace=True)  # 改列名
Dishirb = Dishi_Rb.columns.tolist()  # 把Cols1的列名称，取出来放到一个list里边。即返回['a', 'b', 'c', 'd', 'e', '责任地市']
Dishirb.insert(18, Dishirb.pop(Dishirb.index('故障不过夜处理及时率')))  # pop()把工单开始时间从cols列表里挖出来，通过位置参数“0”，然后放到第一列。
Dishi_Rb = Dishi_Rb[Dishirb]  # 排列后数据。

# -------------------------表3-区县表

Zd_Qx.归档总量 = Zd_Qx.归档总量.fillna('')
# ----------------区县分公司

Quxian = Zd_Qx[['地市']]
Quxian['区县分公司'] = Zd_Qx[['区县分公司']]

# -------当日完成

Drwc = copy.deepcopy(Quxian)
Zrqu = Gd_Zb.groupby(['责任区县']).size().reset_index(name='当日完成')
Zrqu.rename(columns={'责任区县': '区县分公司'}, inplace=True)  # 改列名
Drwc = pd.merge(Drwc, Zrqu, on=['区县分公司'], how='left')  # 拼接
Drwc.当日完成 = Drwc.当日完成.fillna(0)
print('#-------当日完成已完成')

# -------当日投诉派单

Drtspd = copy.deepcopy(Quxian)
Zrqu1 = Gd_Pb.groupby(['责任区县']).size().reset_index(name='当日投诉派单')
Zrqu1.rename(columns={'责任区县': '区县分公司'}, inplace=True)  # 改列名
Drtspd = pd.merge(Drtspd, Zrqu1, on=['区县分公司'], how='left')  # 拼接
Drtspd.当日投诉派单 = Drtspd.当日投诉派单.fillna(0)
print('当日投诉派单已完成')

# ----------------24点在途

Essdzt = copy.deepcopy(Quxian)
Zrqu3 = Gd_Sj[(Gd_Sj.是否要剔除 == False)].groupby(['责任区县']).size().reset_index(name='二十四点在途')
Zrqu3.rename(columns={'责任区县': '区县分公司'}, inplace=True)  # 改列名
Essdzt = pd.merge(Essdzt, Zrqu3, on=['区县分公司'], how='left')  # 拼接
Essdzt.二十四点在途 = Essdzt.二十四点在途.fillna(0)
print('24点在途已完成')

# ----------------压单比小于0.8

Ydbxyldb = Quxian[['地市', '区县分公司']]
Ydbxyldb['压单比小于08'] = (Essdzt.二十四点在途 / Zd_Qx.日均归档量).round(decimals=2)  # 保留两位小数
print('压单比小于0.8已完成')

# -------------在途目标

Ztmb = Quxian[['地市', '区县分公司']]
Ztmb['在途目标'] = Zd_Qx.日均归档量 * 0.8
Ztmb.在途目标 = Ztmb.在途目标 // 1
print('在途目标已完成')


# -----------超长超时工单

def countytime(time2, timebar):
    overwork = copy.deepcopy(Quxian)
    dutycounty = Gd_Sj[(Gd_Sj.是否要剔除 == False) & (Gd_Sj.时间差 > time2)].groupby(['责任区县']).size().reset_index(name='超过N时间')
    dutycounty.rename(columns={'责任区县': '区县分公司'}, inplace=True)  # 改列名
    overwork = pd.merge(overwork, dutycounty, on=['区县分公司'], how='left')  # 拼接
    overwork.超过N时间 = overwork.超过N时间.fillna(0)
    overwork.rename(columns={'超过N时间': timebar}, inplace=True)  # 改列名
    return overwork


dysbxs = countytime(48, '大于四十八小时')
Dyqt = countytime(168, '大于7天')
Dyswt = countytime(360, '大于15天')
Dysst = countytime(720, '大于30天')

# --------超48小时占比

Csbxszb = Quxian[['地市', '区县分公司']]
Csbxszb['超48小时占比'] = (dysbxs.大于四十八小时 / Essdzt.二十四点在途).apply(lambda x: '%.2f%%' % (x * 100))
Csbxszb = Csbxszb.drop(['区县分公司'], axis=1)  # 删除列
print('#--------超48小时占比已完成')

# ----------联接区县表

Quxianbiao = Zd_Qx[['地市', '区县分公司']]
Quxianbiao['当日派单'] = Drtspd[['当日投诉派单']]
Quxianbiao['当日归档'] = Drwc[['当日完成']]
Quxianbiao['二十四点在途'] = Essdzt[['二十四点在途']]
Quxianbiao['压单比小于0.8'] = Ydbxyldb[['压单比小于08']]
Quxianbiao['在途目标'] = Ztmb[['在途目标']]
Quxianbiao['区县分公司1'] = Zd_Qx[['区县分公司']]
Quxianbiao['大于四十八小时'] = dysbxs[['大于四十八小时']]
Quxianbiao['大于7天'] = Dyqt[['大于7天']]
Quxianbiao['大于15天'] = Dyswt[['大于15天']]
Quxianbiao['大于30天'] = Dysst[['大于30天']]
Quxianbiao['超48小时占比'] = Csbxszb[['超48小时占比']]
Quxianbiao['归档总量'] = Zd_Qx[['归档总量']]
Quxianbiao['日均归档量'] = Zd_Qx[['日均归档量']].round(decimals=1)  # 保留两位小数
Quxianbiao.区县分公司 = Quxianbiao.区县分公司.fillna('')
Quxianbiao.区县分公司1 = Quxianbiao.区县分公司1.fillna(0)
print('区县表已完成')

# --------------------表4网格

Zd_Wg.日期 = Zd_Wg.日期.fillna('')
Zd_Wg.归档总量 = Zd_Wg.归档总量.fillna('')
Wangge = Zd_Wg[['地市', '区县分公司', '所属网格', '昨日在途']]

# ------------------24点在途

Wg_Essdzt1 = Gd_Sj[(Gd_Sj.是否要剔除 == False)].groupby(['网格名称']).size().reset_index(name='二十四点在途')
Wg_Essdzt1.rename(columns={'网格名称': '所属网格'}, inplace=True)  # 改列名
Wg_Essdzt1 = pd.merge(Wangge, Wg_Essdzt1, on=['所属网格'], how='left')  # 拼接
Wg_Essdzt1.二十四点在途 = Wg_Essdzt1.二十四点在途.fillna(0)  # 把nan填充为0
print('24点在途已完成')

# -------------------压单比小于等于0.8

Wg_Ydbxyldb = Zd_Wg[['地市', '区县分公司', '所属网格', '昨日在途']]
Wg_Ydbxyldb['压单比小于等于零点八'] = (Wg_Essdzt1.二十四点在途 / Zd_Wg.日均归档量).round(decimals=2)  # 保留两位小数
Wg_Ydbxyldb.压单比小于等于零点八 = Wg_Ydbxyldb.压单比小于等于零点八.fillna(0)  # 把nan填充为0
print('压单比小于等于0.8已完成')

# ----------------------在途目标

Wg_Ztmb = Zd_Wg[['地市', '区县分公司', '所属网格', '昨日在途']]
Wg_Ztmb['在途目标'] = Zd_Wg.日均归档量 * 0.8
Wg_Ztmb.在途目标 = Wg_Ztmb.在途目标 // 1  # 取整数部分
print('在途目标已完成')


# ————————————超长超时工单

def gridovertime(time3, timeline):
    wg_overtime = Gd_Sj[(Gd_Sj.是否要剔除 == False) & (Gd_Sj.时间差 > time3)].groupby(['网格名称']).size().reset_index(
        name='超长超时工单')
    wg_overtime.rename(columns={'网格名称': '所属网格'}, inplace=True)  # 改列名
    wg_overtime = pd.merge(Wangge, wg_overtime, on=['所属网格'], how='left')  # 拼接
    wg_overtime.超长超时工单 = wg_overtime.超长超时工单.fillna(0)  # 把nan填充为0
    wg_overtime.rename(columns={'超长超时工单': timeline}, inplace=True)  # 改列名
    return wg_overtime


Wg_dysbxs = gridovertime(48, '大于四十八小时')
Wg_dyqt = gridovertime(168, '大于7天')
Wg_dyswt = gridovertime(360, '大于15天')
Wg_dysst = gridovertime(720, '大于30天')

# --------------------超48小时占比

Wg_Csbxszb = Zd_Wg[['地市', '区县分公司', '所属网格', '昨日在途']]
Wg_Csbxszb['超48小时占比'] = (Wg_dysbxs.大于四十八小时 / Wg_Essdzt1.二十四点在途).apply(lambda x: '%.2f%%' % (x * 100))
print('超48小时占比已完成')

# ----------联接区县表

Wanggebiao = Zd_Wg[['地市', '区县分公司', '所属网格', '昨日在途']]
Wanggebiao['二十四点在途'] = Wg_Essdzt1[['二十四点在途']]
Wanggebiao['压单比小于等于零点八'] = Wg_Ydbxyldb[['压单比小于等于零点八']]
Wanggebiao['在途目标'] = Wg_Ztmb[['在途目标']]
Wanggebiao['所属网格1'] = Zd_Wg[['所属网格']]
Wanggebiao['大于四十八小时'] = Wg_dysbxs.大于四十八小时
Wanggebiao['大于7天'] = Wg_dyqt.大于7天
Wanggebiao['大于15天'] = Wg_dyswt.大于15天
Wanggebiao['大于30天'] = Wg_dysst.大于30天
Wanggebiao['超48小时占比'] = Wg_Csbxszb[['超48小时占比']]
Wanggebiao['归档总量'] = Zd_Wg[['归档总量']]
Wanggebiao['日均归档量'] = Zd_Wg[['日均归档量']].round(decimals=1)  # 保留两位小数

print('区县表已完成')

# --------------------修改列名

Gd_Sj.rename(columns={'区域类型_x': '区域类型'}, inplace=True)  # 改列名

# 修改当日归档明细

Gd_Zb.rename(columns={'是否超过二十点': todaytime1, '是否超过二十四点': zt_time1}, inplace=True)

# 修改派单明细列表

Gd_Pb.rename(columns={'是否20点到达': todaytime1}, inplace=True)

# 修改在途明细列名

Gd_Sj.rename(columns={'是否超过二十点整': todaytime1, '是否超过二十四点整': zt_time1, '时间差': Sz_Time1}, inplace=True)

# ----------------修改地市日报列名

Dishi_Rb.rename(columns={'二十四点在途': '24点在途'}, inplace=True)  # 改列名
Dishi_Rb.rename(columns={'地市1_x': '地市'}, inplace=True)  # 改列名
Dishi_Rb.rename(columns={'超过四十八小时': '>48小时'}, inplace=True)  # 改列名
Dishi_Rb.rename(columns={'超过七日': '>7天'}, inplace=True)  # 改列名
Dishi_Rb.rename(columns={'超过十五日': '>15天'}, inplace=True)  # 改列名
Dishi_Rb.rename(columns={'超四十八小时占比': '超48小时占比＜15%'}, inplace=True)  # 改列名
Dishi_Rb.rename(columns={'地市1_y': '地市'}, inplace=True)  # 改列名

# ------------------修改区县列名

Quxianbiao.rename(columns={'二十四点在途': '24点在途'}, inplace=True)  # 改列名
Quxianbiao.rename(columns={'压单比小于0.8': '压单比≤0.8'}, inplace=True)  # 改列名
Quxianbiao.rename(columns={'区县分公司1': '区县分公司'}, inplace=True)  # 改列名
Quxianbiao.rename(columns={'大于四十八小时': '>48小时'}, inplace=True)  # 改列名
Quxianbiao.rename(columns={'大于7天': '>7天'}, inplace=True)  # 改列名
Quxianbiao.rename(columns={'大于15天': '>15天'}, inplace=True)  # 改列名
Quxianbiao.rename(columns={'大于30天': '>30日'}, inplace=True)  # 改列名

# ----------------修改网格表

Wanggebiao.rename(columns={'二十四点在途': '24点在途'}, inplace=True)  # 改列名
Wanggebiao.rename(columns={'压单比小于等于零点八': '压单比≤0.8'}, inplace=True)  # 改列名
# Wanggebiao.rename(columns={'八月拍照未归档':'8月拍照未归档'}, inplace=True)  # 改列名
Wanggebiao.rename(columns={'所属网格1': '所属网格'}, inplace=True)  # 改列名
Wanggebiao.rename(columns={'大于四十八小时': '>48小时'}, inplace=True)  # 改列名
Wanggebiao.rename(columns={'大于7天': '>7天'}, inplace=True)  # 改列名
Wanggebiao.rename(columns={'大于30天': '>30日'}, inplace=True)  # 改列名

# 故障累计表的表头名

guzhangname = pd.DataFrame({'故障不过夜-20点前派单': ['']})
zaituname = pd.DataFrame({'故障不过夜-截止当日24点仍在途': ['']})
yueleiname = pd.DataFrame({'故障不过夜-处理及时率': ['']})
# --------------------导出数据

print('导出数据')
with pd.ExcelWriter(Time + '广西移动家宽投诉在途工单.xlsx') as writer:  # 写入结果为当前路径生成Excle表格文件
    Dishi_Rb.to_excel(writer, sheet_name='地市日报', startcol=0, startrow=2, index=False, header=True)
    Gd_Zb.to_excel(writer, sheet_name='当日归档明细', startcol=0, index=False, header=True)
    Gd_Pb.to_excel(writer, sheet_name='当日派单明细', startcol=0, index=False, header=True)
    Gd_Sj.to_excel(writer, sheet_name='当日在途明细', startcol=0, index=False, header=True)
    Quxianbiao.to_excel(writer, sheet_name='表3-区县', startcol=0, index=False, header=True)
    Wanggebiao.to_excel(writer, sheet_name='表4-网格', startcol=0, index=False, header=True)
    guzhangpd0.to_excel(writer, sheet_name='故障不过夜累计', startcol=0, startrow=2, index=False, header=True)
    guzhangzt0.to_excel(writer, sheet_name='故障不过夜累计', startcol=0, startrow=20, index=False, header=True)
    guzhangjsl0.to_excel(writer, sheet_name='故障不过夜累计', startcol=0, startrow=38, index=False, header=True)
    guzhangname.to_excel(writer, sheet_name='故障不过夜累计', startcol=0, startrow=1, index=False, header=True)
    zaituname.to_excel(writer, sheet_name='故障不过夜累计', startcol=0, startrow=19, index=False, header=True)
    yueleiname.to_excel(writer, sheet_name='故障不过夜累计', startcol=0, startrow=37, index=False, header=True)
print('合并单元格')

# 地市日报样式调整

path1 = ('./' + Time + '广西移动家宽投诉在途工单.xlsx')
wb = load_workbook(path1)
sheet = wb.get_active_sheet()
# 合并单元格   
sheet.merge_cells('A1:O1')
sheet['A1'] = '家宽投诉在途工单通报-' + yestday0 + '-截至24时'
sheet.merge_cells('A2:A3')
sheet['A2'] = '地市'
sheet.merge_cells('B2:B3')
sheet['B2'] = '故障不过夜20点前派单'
sheet.merge_cells('C2:C3')
sheet['C2'] = '故障不过夜在途'
sheet.merge_cells('D2:D3')
sheet['D2'] = '当日20~24点归档'
sheet.merge_cells('E2:E3')
sheet['E2'] = '当日派单'
sheet.merge_cells('F2:F3')
sheet['F2'] = '当日归档'
sheet.merge_cells('G2:G3')
sheet['G2'] = '24点在途'
sheet.merge_cells('H2:H3')
sheet['H2'] = '压单比≤0.8'
sheet.merge_cells('I2:I3')
sheet['I2'] = '在途目标'
sheet.merge_cells('J2:J3')
sheet['J2'] = '地市'
sheet.merge_cells('K2:M2')
sheet['K2'] = '超长超时工单'
sheet.merge_cells('N2:N3')
sheet['N2'] = '超48小时占比＜15%'
sheet.merge_cells('O2:O3')
sheet['O2'] = '地市'
sheet.merge_cells('P2:P3')
sheet['P2'] = '全流程处理时长≤16'
sheet.merge_cells('Q2:Q3')
sheet['Q2'] = '30分钟投诉首响及时率≥95%'
sheet.merge_cells('R2:R3')
sheet['R2'] = '全流程处理及时率≥93%'
sheet.merge_cells('S2:S3')
sheet['S2'] = '故障不过夜处理及时率'
print('合并单元格完成')
# 改变表格的字体
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        sheet.cell(row=i, column=j).font = Font(name='微软雅黑', size=11, color="000000", b=False, i=False)


# 给表头字体加粗

def boldfont(row3, row4, column3, column4):
    for i in range(row3, row4):
        for j in range(column3, column4):
            sheet.cell(row=i, column=j).font = Font(name='微软雅黑', size=11, color="000000", b=True, i=False)


boldfont(1, 4, 1, 20)
boldfont(sheet.max_row, sheet.max_row + 1, 1, 20)


# 改字体颜色及大小加粗

def fontcolor(row1, row2, column1, column2, size1, color1, blod):
    for i in range(row1, row2):
        for j in range(column1, column2):
            sheet.cell(row=i, column=j).font = Font(name='微软雅黑', size=size1, color=color1, b=blod, i=False)


fontcolor(1, 2, 1, 2, 14, '000000', True)
fontcolor(2, 3, 11, 12, 10, '000000', True)
fontcolor(2, 3, 2, 5, 11, 'FFFFFF', True)
fontcolor(3, 4, 11, 14, 10, '000000', True)
fontcolor(4, sheet.max_row + 1, 9, 10, 12, '0070C0', False)

# 设置基础颜色
fille = PatternFill("solid", fgColor="FCE4D6")  # 粉红色
fille1 = PatternFill("solid", fgColor="F4B084")  # 压单比色等相同颜色
fille2 = PatternFill("solid", fgColor="92D050")  # 浅绿色
fille3 = PatternFill("solid", fgColor="FFC000")  # 橙色
fille4 = PatternFill("solid", fgColor="FF0000")  # 红色
fille5 = PatternFill("solid", fgColor="F5E59C")  # 浅橙色
fille6 = PatternFill("solid", fgColor="00B0F0")  # 蓝色
fille7 = PatternFill("solid", fgColor="FF9999")  # 粉红色

deeporange = [(2, 8), (2, 14), (2, 16), (2, 17), (2, 18), (2, 19)]
pink = [(2, 1), (2, 5), (2, 6), (2, 7), (2, 9), (2, 10), (2, 11), (2, 15), (3, 11), (3, 12), (3, 13)]
blue = [(2, 2), (2, 3), (2, 4)]


def fillcolor(filename, color):
    for i, element1 in enumerate(filename):
        sheet.cell(row=element1[0], column=element1[1]).fill = color


fillcolor(deeporange, fille1)
fillcolor(pink, fille)
fillcolor(blue, fille6)
# 压单比≤0.8
for i in range(4, sheet.max_row + 1):
    ydb = sheet.cell(row=i, column=8).value
    if ydb <= 0.8:
        sheet.cell(row=i, column=8).fill = fille2
    elif 0.8 < ydb <= 1:
        sheet.cell(row=i, column=8).fill = fille5
    elif 1 < ydb <= 1.2:
        sheet.cell(row=i, column=8).fill = fille3
    else:
        sheet.cell(row=i, column=8).fill = fille4

# 超48小时占比<15%

for i in range(4, sheet.max_row + 1):
    csb = sheet.cell(row=i, column=14).value
    csb = float(csb.strip('%'))
    if csb < 15:
        sheet.cell(row=i, column=14).fill = fille2
    elif 15 <= csb <= 20:
        sheet.cell(row=i, column=14).fill = fille5
    elif 20 < csb <= 25:
        sheet.cell(row=i, column=14).fill = fille3
    else:
        sheet.cell(row=i, column=14).fill = fille4

# 全流程处理时长≤16
for i in range(4, sheet.max_row + 1):
    qlc = sheet.cell(row=i, column=16).value

    if qlc <= 16:
        sheet.cell(row=i, column=16).fill = fille2
    elif 16 < qlc <= 17:
        sheet.cell(row=i, column=16).fill = fille5
    elif 17 < qlc <= 20:
        sheet.cell(row=i, column=16).fill = fille3
    else:
        sheet.cell(row=i, column=16).fill = fille4
    # 30分钟投诉首响及时率≥95%
for i in range(4, sheet.max_row + 1):
    tsl = sheet.cell(row=i, column=17).value
    tsl = float(tsl.strip('%'))
    if tsl >= 95:
        sheet.cell(row=i, column=17).fill = fille2
    elif 90 <= tsl < 95:
        sheet.cell(row=i, column=17).fill = fille5
    elif 82 <= tsl < 90:
        sheet.cell(row=i, column=17).fill = fille3
    else:
        sheet.cell(row=i, column=17).fill = fille4
# 全流程处理及时率≥93%

for i in range(4, sheet.max_row + 1):
    qcl = sheet.cell(row=i, column=18).value
    qcl = float(qcl.strip('%'))
    if qcl >= 93:
        sheet.cell(row=i, column=18).fill = fille2
    elif 90 <= qcl < 93:
        sheet.cell(row=i, column=18).fill = fille5
    elif 85 <= qcl < 90:
        sheet.cell(row=i, column=18).fill = fille3
    else:
        sheet.cell(row=i, column=18).fill = fille4
# 故障不过夜处理及时率

hsmlist = list()
for i in range(4, sheet.max_row + 1):
    yuelj = sheet.cell(row=i, column=19).value
    yuelj = float(yuelj.strip('%'))
    hsmlist.append(yuelj)
hsmlist.sort(key=None, reverse=False)
for i in range(4, sheet.max_row + 1):
    yuelil = sheet.cell(row=i, column=19).value
    yuelil = float(yuelil.strip('%'))
    if yuelil <= hsmlist[2]:
        sheet.cell(row=i, column=19).fill = fille7
    # 字体对齐方式
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
        sheet.cell(row=i, column=j).alignment = alignment

    # 边框样式

for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        side_border = Side(style="thin", color="000000")  # 黑色色
        border = Border(left=side_border, right=side_border, top=side_border, bottom=side_border)
        sheet.cell(row=i, column=j).border = border
print('填充颜色完成')
path2 = (Time + '广西移动家宽投诉在途工单.xlsx')
wb.save(path2)
print('已全部完成，结束程序')
